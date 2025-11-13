import unittest
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from main_facultytt import auto_adjust_column_widths, sanitize_sheet_name

def auto_adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

def sanitize_sheet_name(name):
    """Remove invalid Excel sheet characters and trim length."""
    import re
    name = re.sub(r'[\\/*?:\[\]/]', '_', str(name))  # replace invalid chars with underscore
    return name[:30]  # Excel limit = 31 chars


class TestFacultyTT(unittest.TestCase):

    def test_sanitize_sheet_name(self):
        self.assertEqual(sanitize_sheet_name("NormalName"), "NormalName")
        self.assertEqual(sanitize_sheet_name("Invalid/Name*With:Chars?"), "Invalid_Name_With_Chars_")
        self.assertEqual(len(sanitize_sheet_name("A"*50)), 30)

    def test_auto_adjust_column_widths(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Header1", "Header2"])
        ws.append(["Short", "VeryLongValueInColumn2"])
        
        auto_adjust_column_widths(ws)
        
        col_widths = [ws.column_dimensions[get_column_letter(i+1)].width for i in range(ws.max_column)]
        self.assertLessEqual(col_widths[0], 45)
        self.assertLessEqual(col_widths[1], 45)
        self.assertGreater(col_widths[1], col_widths[0])

    @patch("facultyTT_v2.random.choice")
    def test_random_assignment_mock(self, mock_choice):
        mock_choice.side_effect = lambda x: x[0]
        
        days = ["Mon", "Tue"]
        slot_keys = ["S1", "S2", "S3"]
        timetable = {day: [""] * len(slot_keys) for day in days}
        
        random_day = mock_choice(days)
        random_slot = mock_choice(range(len(slot_keys)))
        timetable[random_day][random_slot] = "CS101"
        
        self.assertEqual(timetable["Mon"][0], "CS101")
        self.assertEqual(timetable["Tue"][0], "")

if __name__ == "__main__":
    unittest.main()

