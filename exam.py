# exam_scheduler_final.py
import re, math, os
import pandas as pd
from collections import defaultdict, deque
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.cell.cell import MergedCell

# ---- Config ----
SLOT_LABELS = ["09:00-12:00", "14:00-17:00"]
MAX_GLOBAL_EXAMS_PER_DAY = 4
MAX_EXAMS_PER_GROUP_PER_DAY = 1
DEFAULT_START_DATE = "2025-11-20"
ROOM_SORT_MODE = "small-first"
USE_HALLS_LAST = True

def invigilators_needed(capacity):
    return 2 if capacity >= 200 else 1

def extract_semester_id(group_name: str) -> str:
    m = re.search(r"(\d+)", str(group_name))
    return m.group(1) if m else str(group_name)

class Course:
    def __init__(self, row, group_name):
        self.group = group_name
        self.code = str(row["Course_Code"]).strip()
        self.title = str(row.get("Course_Title", self.code)).strip()
        try:
            self.students = int(str(row.get("Students", "0")).strip())
        except:
            self.students = 0
        flag = str(row.get("Elective", "0")).strip()
        self.is_elective = flag.lower() in ("1", "true", "yes")

class ExamScheduler:
    def __init__(self, rooms_file, departments, faculty_file, students_file, start_date=DEFAULT_START_DATE):
        self.rooms_df = pd.read_csv(rooms_file)
        self.departments = departments
        self.invig_df = pd.read_csv(faculty_file)
        self.students_df = pd.read_csv(students_file)
        self.start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        self.groups = list(departments.keys())
        self.invigilators = sorted([str(n).strip() for n in self.invig_df["Name"] if str(n).strip()])

        self.rooms = self._load_rooms()
        self.room_by_id = {r["Room_ID"]: r for r in self.rooms}
        self.courses = self._load_courses()

        self.room_remaining = {}
        self.group_daily = {}
        self.global_daily = {}
        self.used_rooms = {}

        self.scheduled = []
        self.unscheduled = []
        self.invig_assignments = []
        self._inv_idx = 0

        # Build student pools from students.csv (Option B: Courses semicolon-separated)
        # pools_course[(group, course)] -> deque(student_ids)
        # group_pool[group] -> deque(student_ids) fallback
        self.pools_course = defaultdict(deque)
        self.group_pool = defaultdict(deque)
        for _, r in self.students_df.iterrows():
            sid = str(r.get("Student_ID", "")).strip()
            grp = str(r.get("Group", "")).strip()
            courses_raw = str(r.get("Courses", "")).strip()
            if not sid or not grp:
                continue
            courses = [c.strip() for c in courses_raw.split(";") if c.strip()]
            for c in courses:
                self.pools_course[(grp, c)].append(sid)
            # Also put in group fallback
            self.group_pool[grp].append(sid)

    def _load_rooms(self):
        rooms = []
        for _, r in self.rooms_df.iterrows():
            rid = str(r["Room_ID"]).strip()
            t = str(r["Type"]).strip().lower()
            try:
                cap = int(str(r["Capacity"]).strip())
            except:
                cap = 0
            if cap <= 0:
                continue
            is_lab = ("lab" in t) or rid.upper().startswith(("L", "H"))
            is_library = ("library" in t)
            is_hall = ("hall" in t) or rid.upper() in {"C002", "C003", "C004"}
            if is_lab or is_library:
                continue
            usable = math.ceil(cap / 2)
            rooms.append({"Room_ID": rid, "Type": t, "Capacity": cap, "Usable": usable, "IsHall": bool(is_hall)})
        rooms.sort(key=lambda x: (x["Usable"], x["Room_ID"]))
        return rooms

    def _load_courses(self):
        out = {}
        for g, file in self.departments.items():
            df = pd.read_csv(file)
            lst = [Course(r, g) for _, r in df.iterrows() if int(str(r.get("Students", 0))) > 0]
            lst.sort(key=lambda c: (-c.students, c.code))
            out[g] = lst
        return out

    def _ensure_date(self, date):
        if date not in self.room_remaining:
            self.room_remaining[date] = {s: {r["Room_ID"]: r["Usable"] for r in self.rooms} for s in SLOT_LABELS}
        if date not in self.used_rooms:
            self.used_rooms[date] = {s: set() for s in SLOT_LABELS}
        if date not in self.group_daily:
            self.group_daily[date] = {g: 0 for g in self.groups}
        if date not in self.global_daily:
            self.global_daily[date] = 0

    def _ordered(self, ids, remaining):
        if ROOM_SORT_MODE == "small-first":
            return sorted(ids, key=lambda rid: (remaining.get(rid, 0), rid))
        else:
            return sorted(ids, key=lambda rid: (-remaining.get(rid, 0), rid))

    def _alloc_rooms(self, date, slot, need):
        remaining = self.room_remaining[date][slot]
        normal_ids = [r["Room_ID"] for r in self.rooms if not r["IsHall"]]
        hall_ids = [r["Room_ID"] for r in self.rooms if r["IsHall"]]
        def try_allocate(candidates):
            alloc, total = [], 0
            for rid in self._ordered(candidates, remaining):
                avail = remaining.get(rid, 0)
                usable_cap = self.room_by_id[rid]["Usable"]
                avail = min(avail, usable_cap)
                if avail <= 0:
                    continue
                take = min(avail, need - total)
                if take > 0:
                    alloc.append((rid, take))
                    total += take
                if total >= need:
                    break
            return alloc if total >= need else None
        if USE_HALLS_LAST:
            alloc = try_allocate([rid for rid in normal_ids if remaining.get(rid, 0) > 0])
            if alloc is not None:
                return alloc
            return try_allocate([rid for rid in normal_ids + hall_ids if remaining.get(rid, 0) > 0])
        else:
            all_ids = [rid for rid in normal_ids + hall_ids if remaining.get(rid, 0) > 0]
            return try_allocate(all_ids)

    def _book_alloc(self, date, slot, alloc):
        for rid, cnt in alloc:
            usable_cap = self.room_by_id[rid]["Usable"]
            safe_cnt = min(cnt, usable_cap)
            self.room_remaining[date][slot][rid] -= safe_cnt
            self.used_rooms[date][slot].add(rid)

    def _place_merged_course(self, code, title, students, groups_set, date, slot):
        if self.global_daily[date] >= MAX_GLOBAL_EXAMS_PER_DAY:
            return False
        for g in groups_set:
            if self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY:
                return False
        alloc = self._alloc_rooms(date, slot, students)
        if alloc is None:
            return False
        sanitized = [(rid, min(cnt, self.room_by_id[rid]["Usable"])) for rid, cnt in alloc]
        self._book_alloc(date, slot, sanitized)
        for g in groups_set:
            self.group_daily[date][g] += 1
        self.global_daily[date] += 1
        alloc_text = "; ".join([f"{rid}:{cnt}" for rid, cnt in sanitized])
        self.scheduled.append({"Date": date.strftime("%Y-%m-%d"), "Slot": slot, "Groups": ", ".join(sorted(groups_set)),
                               "Course_Code": code, "Course_Title": title, "Students": students, "Allocations": alloc_text})
        return True

    def _plan_electives_by_semester(self):
        pool = {}
        for g in self.groups:
            sem = extract_semester_id(g)
            for c in self.courses[g]:
                if c.is_elective:
                    if sem not in pool:
                        pool[sem] = {}
                    if c.code not in pool[sem]:
                        pool[sem][c.code] = {"electives": [], "groups": set()}
                    pool[sem][c.code]["electives"].append(c)
                    pool[sem][c.code]["groups"].add(g)
        return pool

    def _remove_scheduled_electives_from_pool(self):
        for g in self.groups:
            self.courses[g] = [c for c in self.courses[g] if not c.is_elective]

    def generate(self):
        pool = self._plan_electives_by_semester()
        def extract_number(s):
            m = re.search(r"(\d+)", str(s))
            return int(m.group(1)) if m else 999
        semesters = sorted(pool.keys(), key=lambda x: extract_number(x))
        day_cursor = 0
        for sem in semesters:
            course_blocks = pool[sem]
            course_items = sorted(course_blocks.items(), key=lambda kv: kv[0])
            mid = len(course_items) // 2
            morning_items = course_items[:mid]
            afternoon_items = course_items[mid:]
            date = self.start_date + timedelta(days=day_cursor)
            self._ensure_date(date)
            for slot, items in zip([SLOT_LABELS[0], SLOT_LABELS[1]], [morning_items, afternoon_items]):
                for code, block in items:
                    total_students = sum(c.students for c in block["electives"])
                    alloc = self._alloc_rooms(date, slot, total_students)
                    if alloc is None:
                        continue
                    sanitized = [(rid, min(cnt, self.room_by_id[rid]["Usable"])) for rid, cnt in alloc]
                    self._book_alloc(date, slot, sanitized)
                    groups_set = block["groups"]
                    alloc_text = "; ".join(f"{rid}:{cnt}" for rid, cnt in sanitized)
                    self.scheduled.append({"Date": date.strftime("%Y-%m-%d"), "Slot": slot,
                                           "Groups": ", ".join(sorted(groups_set)),
                                           "Course_Code": code, "Course_Title": block["electives"][0].title,
                                           "Students": total_students, "Allocations": alloc_text})
            day_cursor += 1
        self._remove_scheduled_electives_from_pool()

        merged_regular = {}
        for g in self.groups:
            for c in self.courses[g]:
                if c.is_elective:
                    continue
                if c.code not in merged_regular:
                    merged_regular[c.code] = {"code": c.code, "title": c.title or c.code, "students": 0, "groups": set()}
                merged_regular[c.code]["students"] += c.students
                merged_regular[c.code]["groups"].add(g)
        pending = sorted(merged_regular.values(), key=lambda x: (-x["students"], x["code"]))
        day = 0
        while pending and day < 300:
            date = self.start_date + timedelta(days=day)
            self._ensure_date(date)
            placed_today = 0
            si = 0
            i = 0
            while i < len(pending) and placed_today < MAX_GLOBAL_EXAMS_PER_DAY:
                exam = pending[i]
                if any(self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY for g in exam["groups"]):
                    i += 1
                    continue
                slot = SLOT_LABELS[si % len(SLOT_LABELS)]
                ok = self._place_merged_course(exam["code"], exam["title"], exam["students"], exam["groups"], date, slot)
                if ok:
                    pending.pop(i)
                    placed_today += 1
                    si += 1
                else:
                    i += 1
            day += 1
        if pending:
            for exam in pending:
                self.unscheduled.append({"Group": ", ".join(sorted(exam["groups"])), "Course_Code": exam["code"],
                                         "Course_Title": exam["title"], "Students": exam["students"]})
        self._assign_invigilators()

    def _assign_invigilators(self):
        for d in sorted(self.used_rooms.keys()):
            assigned_today = set()
            for slot in SLOT_LABELS:
                rooms = sorted(list(self.used_rooms[d][slot]))
                for rid in rooms:
                    cap = self.room_by_id[rid]["Capacity"]
                    k = invigilators_needed(cap)
                    picks = []
                    tries = 0
                    while len(picks) < k and self.invigilators and tries < len(self.invigilators)*2:
                        name = self.invigilators[self._inv_idx % len(self.invigilators)]
                        self._inv_idx += 1
                        tries += 1
                        if name not in assigned_today:
                            picks.append(name)
                            assigned_today.add(name)
                    date_str = d.strftime("%Y-%m-%d")
                    exam_names = []
                    for rec in self.scheduled:
                        if rec["Date"] == date_str and rec["Slot"] == slot:
                            for a in rec["Allocations"].split(";"):
                                if ":" in a:
                                    rid2 = a.split(":")[0].strip()
                                    if rid2 == rid:
                                        exam_names.append(rec["Course_Code"])
                                        break
                    self.invig_assignments.append({"Date": date_str, "Slot": slot, "Room_ID": rid,
                                                   "Exam": " | ".join(sorted(set(exam_names))),
                                                   "Invigilators": ", ".join(picks)})

    # Seating assignment: use pools_course[(group, course)] first, then group fallback
    def _assign_students_to_room_alloc(self):
        assigned = defaultdict(list)
        # shallow copies of deques so function doesn't exhaust original pools in scheduler object
        pools_course = {k: deque(v) for k, v in self.pools_course.items()}
        group_pool = {k: deque(v) for k, v in self.group_pool.items()}

        for rec in self.scheduled:
            date = rec["Date"]
            slot = rec["Slot"]
            code = rec["Course_Code"]
            groups = [g.strip() for g in str(rec.get("Groups", "")).split(",") if g.strip()]
            parts = [p.strip() for p in str(rec.get("Allocations", "")).split(";") if p.strip()]

            # prepare round-robin group queue
            grp_queue = deque([g for g in groups if (g, code) in pools_course or g in group_pool])
            if not grp_queue:
                # fallback: if no group-specific pools, take any group in scheduler groups
                grp_queue = deque([g for g in groups])

            for part in parts:
                if ":" not in part:
                    continue
                rid, cnts = part.split(":")
                rid = rid.strip()
                needed = int(cnts.strip())
                picks = []
                while needed > 0 and any((pools_course.get((g, code)) and pools_course[(g, code)]) or (group_pool.get(g) and group_pool[g]) for g in groups):
                    # cycle groups to find available student
                    chosen = None
                    for _ in range(len(grp_queue)):
                        g = grp_queue[0]
                        grp_queue.rotate(-1)
                        if pools_course.get((g, code)) and pools_course[(g, code)]:
                            chosen = ('course', g)
                            break
                        if group_pool.get(g) and group_pool[g]:
                            chosen = ('group', g)
                            break
                    if chosen is None:
                        break
                    typ, g = chosen
                    if typ == 'course':
                        sid = pools_course[(g, code)].popleft()
                    else:
                        sid = group_pool[g].popleft()
                    picks.append(sid)
                    needed -= 1
                # if still needed, try any other course-specific pools
                other_keys = [k for k in pools_course.keys() if k[1] == code and pools_course[k]]
                while needed > 0 and other_keys:
                    k = other_keys.pop(0)
                    sid = pools_course[k].popleft()
                    picks.append(sid)
                    needed -= 1
                # placeholders as empty strings (so Excel cell blank)
                while needed > 0:
                    picks.append("")
                    needed -= 1
                for sid in picks:
                    assigned[(rid, date, slot)].append((sid, code))
        return assigned

    def _place_in_room_grid(self, student_list):
        # student_list: list of (student_id, exam_code)
        SEAT_COLS = 8
        SEAT_ROWS = 6
        # bucket by (year, branch) for mixing
        buckets = defaultdict(list)
        for sid, code in student_list:
            s = sid or ""
            year = s[:2]
            branch = s[2:5] if len(s) >= 5 else "XXX"
            buckets[(year, branch)].append((sid, code))
        deques = [deque(v) for v in buckets.values()] if buckets else []
        interleaved = []
        # round-robin across buckets
        while any(deques):
            for q in deques:
                if q:
                    interleaved.append(q.popleft())
        # ensure length exactly 48 (pad with empty)
        while len(interleaved) < SEAT_COLS * SEAT_ROWS:
            interleaved.append(("", ""))
        interleaved = interleaved[:SEAT_COLS * SEAT_ROWS]
        # build grid rows: 6 rows × 8 cols
        grid = []
        idx = 0
        for r in range(SEAT_ROWS):
            row = []
            for c in range(SEAT_COLS):
                sid, code = interleaved[idx]
                row.append(sid)
                idx += 1
            grid.append(row)
        return grid

    def _build_merged(self):
        rows = self.scheduled
        groups = {}
        title_map = {}
        for r in rows:
            k = (r["Date"], r["Slot"], r["Course_Code"])
            title_map[r["Course_Code"]] = r.get("Course_Title", r["Course_Code"])
            if k not in groups:
                groups[k] = {"Students": 0, "Alloc": {}, "Groups": set()}
            groups[k]["Students"] += int(r.get("Students", 0) or 0)
            parts = [p.strip() for p in str(r.get("Allocations", "")).split(";") if p.strip()]
            for p in parts:
                if ":" in p:
                    rid, cnt = p.split(":")
                    rid, cnt = rid.strip(), int(cnt.strip())
                    groups[k]["Alloc"][rid] = groups[k]["Alloc"].get(rid, 0) + cnt
            gs = str(r.get("Groups", "")).strip()
            if gs:
                for gname in [x.strip() for x in gs.split(",") if x.strip()]:
                    groups[k]["Groups"].add(gname)
        merged_rows = []
        for (date, slot, code), v in sorted(groups.items()):
            merged_rows.append({"Date": date, "Slot": slot, "Course_Code": code, "Students": v["Students"],
                                "Allocations": "; ".join(f"{rid}:{cnt}" for rid, cnt in v["Alloc"].items()),
                                "Groups": ", ".join(sorted(v["Groups"]))})
        legend = sorted([(code, title) for code, title in title_map.items()], key=lambda x: x[0])
        return pd.DataFrame(merged_rows), pd.DataFrame(legend, columns=["Course_Code", "Course_Title"])

    def _build_grid(self, merged_df):
        dates = sorted(merged_df["Date"].unique())
        grid = pd.DataFrame(index=dates, columns=SLOT_LABELS)
        for d in dates:
            for s in SLOT_LABELS:
                subset = merged_df[(merged_df["Date"] == d) & (merged_df["Slot"] == s)]
                grid.at[d, s] = ", ".join(subset["Course_Code"].tolist()) if not subset.empty else ""
        return grid

    def export(self, out="final_exam_schedule_with_seating.xlsx"):
        merged_df, legend_df = self._build_merged()
        grid_df = self._build_grid(merged_df)

        room_student_map = self._assign_students_to_room_alloc()
        room_grids = {}
        for (rid, date, slot), students in room_student_map.items():
            grid = self._place_in_room_grid(students)
            room_grids[(rid, date, slot)] = grid

        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            merged_df.to_excel(writer, sheet_name="Exam Schedule", index=False, startrow=1)
            grid_df.to_excel(writer, sheet_name="Exam Schedule", startrow= len(merged_df) + 6)
            legend_df.to_excel(writer, sheet_name="Exam Schedule", index=False, startrow= len(merged_df) + len(grid_df) + 10)

            # invigilators sheet
            invig_rows = []
            for rec in self.invig_assignments:
                date = rec["Date"]; slot = rec["Slot"]; room = rec["Room_ID"]
                invs = [x.strip() for x in rec["Invigilators"].split(",") if x.strip()]
                exam = rec.get("Exam", "")
                for inv in invs:
                    invig_rows.append({"Date": date, "Slot": slot, "Room_ID": room, "Exam": exam, "Invigilator": inv})
            if invig_rows:
                pd.DataFrame(invig_rows).to_excel(writer, sheet_name="Invigilators", index=False)

            # seating summary
            summary = []
            for key, grid in room_grids.items():
                rid, date, slot = key
                count = sum(1 for row in grid for v in row if v)
                summary.append({"Room_ID": rid, "Date": date, "Slot": slot, "Placed": count})
            if summary:
                pd.DataFrame(summary).to_excel(writer, sheet_name="Seating Plans Summary", index=False)

        wb = load_workbook(out)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        thick = Border(left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium"))
        gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        pale_blue = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        pale_pink = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

        # create one seating sheet per date
        by_date = defaultdict(list)
        for (rid, date, slot), grid in room_grids.items():
            by_date[date].append((rid, slot, grid))

        # Remove existing seating sheets named "Seating Plan - "
        for name in [n for n in wb.sheetnames if n.startswith("Seating Plan - ")]:
            wsold = wb[name]
            wb.remove(wsold)

        for date, items in sorted(by_date.items()):
            sheet_name = f"Seating Plan - {date}"
            ws = wb.create_sheet(sheet_name)
            ws.sheet_view.showGridLines = False
            row_cursor = 1
            SEAT_COLS = 8
            SEAT_ROWS = 6

            for rid, slot, grid in items:
                # header row: Room label + BOARD merged
                ws.cell(row=row_cursor, column=2, value="Room").border = thin
                cell_room = ws.cell(row=row_cursor, column=3, value=rid)
                cell_room.alignment = Alignment(horizontal="center", vertical="center")
                cell_room.border = thin
                row_cursor += 1

                ws.merge_cells(start_row=row_cursor, start_column=2, end_row=row_cursor, end_column=SEAT_COLS+1)
                b = ws.cell(row=row_cursor, column=2, value="BOARD")
                b.alignment = Alignment(horizontal="center", vertical="center")
                b.font = Font(bold=True, size=14)
                b.border = thick
                row_cursor += 1

                ws.cell(row=row_cursor, column=SEAT_COLS+3, value="Date").border = thin
                ws.cell(row=row_cursor, column=SEAT_COLS+4, value=date).border = thin
                row_cursor += 1
                ws.cell(row=row_cursor, column=SEAT_COLS+3, value="Session").border = thin
                ws.cell(row=row_cursor, column=SEAT_COLS+4, value=slot).border = thin
                row_cursor += 1

                start_table_row = row_cursor

                # WINDOW sidebar (merge)
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor + SEAT_ROWS - 1, end_column=1)
                w = ws.cell(row=row_cursor, column=1, value="WINDOW")
                w.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")
                w.border = thick

                # DOOR sidebar (after SEAT_COLS + 1 i.e. col 10)
                ws.merge_cells(start_row=row_cursor, start_column=SEAT_COLS+2, end_row=row_cursor + SEAT_ROWS - 1, end_column=SEAT_COLS+2)
                d = ws.cell(row=row_cursor, column=SEAT_COLS+2, value="DOOR")
                d.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")
                d.border = thick

                # headers C1..C8 (cols 2..9)
                headers = [f"C{i+1}" for i in range(SEAT_COLS)]
                for i, h in enumerate(headers):
                    cell = ws.cell(row=row_cursor, column=2+i, value=h)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = gray
                    cell.border = thin
                row_cursor += 1

                # seat rows (6 rows)
                for r0 in range(SEAT_ROWS):
                    ws.cell(row=row_cursor, column=1).border = thick
                    ws.cell(row=row_cursor, column=SEAT_COLS+2).border = thick
                    for c0 in range(SEAT_COLS):
                        val = grid[r0][c0] if r0 < len(grid) and c0 < len(grid[r0]) else ""
                        cell = ws.cell(row=row_cursor, column=2+c0, value=val)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin
                        cell.fill = pale_blue if (r0 % 2 == 0) else pale_pink
                    row_cursor += 1

                end_table_row = row_cursor - 1

                # outer frame (cols 2..9)
                for rr in range(start_table_row, end_table_row + 1):
                    ws.cell(row=rr, column=2).border = Border(left=Side(style="medium"))
                    ws.cell(row=rr, column=SEAT_COLS+1).border = Border(right=Side(style="medium"))
                for cc in range(2, SEAT_COLS+2):
                    ws.cell(row=start_table_row, column=cc).border = thick
                    ws.cell(row=end_table_row, column=cc).border = thick

                row_cursor += 2

        # format Invigilators sheet if exists
        if "Invigilators" in wb.sheetnames:
            ws_inv = wb["Invigilators"]
            for row in ws_inv.iter_rows():
                for c in row:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = thin

        # adjust column widths (skip merged cells)
        for ws in wb.worksheets:
            for col_cells in ws.columns:
                real_cell = None
                for c in col_cells:
                    if not isinstance(c, MergedCell):
                        real_cell = c
                        break
                if real_cell is None:
                    continue
                col_letter = real_cell.column_letter
                max_len = 0
                for cell in col_cells:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.save(out)
        print(f"Exported -> {out}")


def run_example():
    departments = {
        "CSEA-1": "data/coursesCSEA-I.csv",
        "CSEA-3": "data/coursesCSEA-III.csv",
        "CSEB-3": "data/coursesCSEB-III.csv",
        "CSE-V": "data/coursesCSE-V.csv",
        "DSAI-1": "data/coursesDSAI-I.csv",
        "DSAI-3": "data/coursesDSAI-III.csv",
        "DSAI-5": "data/coursesDSAI-V.csv",
        "ECE-1": "data/coursesECE-I.csv",
        "ECE-3": "data/coursesECE-III.csv",
        "ECE-V": "data/coursesECE-V.csv",
        "Sem-7": "data/courses7.csv"
    }
    rooms = "data/rooms.csv"
    faculty = "data/faculty.csv"
    students = "data/students.csv"
    if not os.path.exists(rooms):
        raise FileNotFoundError("rooms.csv not found")
    if not os.path.exists(faculty):
        raise FileNotFoundError("faculty.csv not found")
    if not os.path.exists(students):
        raise FileNotFoundError("students.csv not found")
    s = ExamScheduler(rooms, departments, faculty, students)
    s.generate()
    s.export("final_exam_schedule_with_seating.xlsx")

if __name__ == "__main__":
    run_example()
