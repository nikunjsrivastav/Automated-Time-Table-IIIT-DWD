import pandas as pd
import json
import random
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

with open("time_slots.json") as f:
    slots = json.load(f)["time_slots"]

slot_keys = [f"{slot['start'].strip()}-{slot['end'].strip()}" for slot in slots]
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

csv_files = [f for f in os.listdir() if f.startswith("courses") and f.endswith(".csv")]
all_courses = []

for file in csv_files:
    df = pd.read_csv(file)
    if "Faculty" in df.columns:
        df["BranchFile"] = file
        all_courses.append(df)

if not all_courses:
    raise ValueError("No valid course CSVs found with 'Faculty' column.")

courses = pd.concat(all_courses, ignore_index=True)

faculty_groups = courses.groupby("Faculty")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

colors = [
    "B8CCE4", "D8E4BC", "F2DCDB", "E5B8B7", "CCC0DA",
    "C6D9F0", "DCE6F1", "E6B8B7", "FCD5B4", "FFF2CC",
    "D9EAD3", "C9DAF8", "F4CCCC", "D0E0E3", "FCE5CD"
]
def auto_adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)
def sanitize_sheet_name(name):
    """Remove invalid Excel sheet characters and trim length."""
    name = re.sub(r'[\\/*?:\[\]/]', '_', str(name))
    return name[:30]
wb = Workbook()
wb.remove(wb.active)

for faculty, df in faculty_groups:
    safe_title = sanitize_sheet_name(faculty)
    ws = wb.create_sheet(title=safe_title)
    ws.append(["Day"] + slot_keys)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="A7BFDE", end_color="A7BFDE", fill_type="solid")
    timetable = {day: [""] * len(slot_keys) for day in days}
    for _, row in df.iterrows():
        course_code = row["Course_Code"]
        semester_half = row.get("Semester_Half", 1)
        course_label = f"{course_code} (H{semester_half})"
        random_day = random.choice(days)
        random_slot = random.choice(range(len(slot_keys)))
        timetable[random_day][random_slot] = course_label
    for day in days:
        ws.append([day] + timetable[day])
    available_colors = colors.copy()
    random.shuffle(available_colors)
    course_colors = {}

    for row_idx in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            val = str(cell.value).strip() if cell.value else ""

            if val:
                course_key = val.split()[0]
                if course_key not in course_colors:
                    if available_colors:
                        course_colors[course_key] = available_colors.pop()
                    else:
                        r = lambda: random.randint(160, 255)
                        course_colors[course_key] = f"{r():02X}{r():02X}{r():02X}"

                cell.fill = PatternFill(
                    start_color=course_colors[course_key],
                    end_color=course_colors[course_key],
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(bold=True)

    auto_adjust_column_widths(ws)
wb.save("facultyTT.xlsx")
print("✅ Faculty timetable created successfully as 'facultyTT.xlsx'")
