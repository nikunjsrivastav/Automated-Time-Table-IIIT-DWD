import pandas as pd
import json
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ================== CONFIG ==================
RANDOM_SEED = 123
random.seed(RANDOM_SEED)

MAX_ATTEMPTS = 30
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded_slots = ["07:30-09:00", "13:15-14:00", "17:30-18:30"]

colors = [
    "FFB3BA", "BAE1FF", "BAFFC9", "FFFFBA", "FFD8BA", "E3BAFF", "D0BAFF",
    "FFCBA4", "C7FFD8", "B8E1FF", "F7FFBA", "FFDFBA", "E9BAFF", "BAFFD9",
    "FFE1BA", "BAFFF2", "D1FFBA"
]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# ================== LOAD DATA ==================
with open("data/time_slots.json") as f:
    slots_raw = json.load(f)["time_slots"]

def parse_time(t):
    h, m = map(int, t.split(":"))
    return h * 60 + m

def slot_duration_from_bounds(start, end):
    return (parse_time(end) - parse_time(start)) / 60.0

slots_norm = []
for s in slots_raw:
    start = s["start"].strip()
    end = s["end"].strip()
    key = f"{start}-{end}"
    dur = slot_duration_from_bounds(start, end)
    slots_norm.append({"key": key, "start": start, "end": end, "duration": dur})
slots_norm.sort(key=lambda x: parse_time(x["start"]))

slot_keys = [s["key"] for s in slots_norm]
slot_durations = {s["key"]: s["duration"] for s in slots_norm}

coursesA = pd.read_csv("data/coursesCSEA-III.csv").to_dict(orient="records")
coursesB = pd.read_csv("data/coursesCSEB-III.csv").to_dict(orient="records")
coursesECE = pd.read_csv("data/coursesECE-III.csv").to_dict(orient="records")
coursesDSAI = pd.read_csv("data/coursesDSAI-III.csv").to_dict(orient="records")

rooms_df = pd.read_csv("data/rooms.csv")
rooms_df["Type"] = rooms_df["Type"].astype(str)
classrooms = rooms_df[rooms_df["Type"].str.lower() == "classroom"]["Room_ID"].tolist()
labs = rooms_df[rooms_df["Type"].str.lower() == "lab"]["Room_ID"].tolist()

# ================== HELPER FUNCTIONS ==================
def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    return str(val).strip()

def parse_ltp(sc_string):
    try:
        parts = [x.strip() for x in sc_string.split("-")]
        while len(parts) < 5:
            parts.append("0")
        return list(map(int, parts[:5]))
    except:
        return [0, 0, 0, 0, 0]

def get_free_blocks(timetable, day):
    free_blocks = []
    block = []
    for slot in slot_keys:
        if slot in excluded_slots:
            if block:
                free_blocks.append(block)
                block = []
            continue
        if timetable.at[day, slot] == "":
            block.append(slot)
        else:
            if block:
                free_blocks.append(block)
                block = []
    if block:
        free_blocks.append(block)
    return free_blocks

def allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code,
                     duration_hours, session_type="L", is_elective=False, labs_on_days=set()):
    if session_type == "P" and day in labs_on_days:
        return False

    free_blocks = get_free_blocks(timetable, day)
    for block in free_blocks:
        total = sum(slot_durations[s] for s in block)
        if total + 1e-9 >= duration_hours:
            slots_to_use = []
            dur_accum = 0.0
            for s in block:
                slots_to_use.append(s)
                dur_accum += slot_durations[s]
                if dur_accum + 1e-9 >= duration_hours:
                    break

            if any(s in excluded_slots for s in slots_to_use):
                continue

            if faculty:
                fac_map = lecturer_busy.get(day, {})
                occupied = fac_map.get(faculty, set())
                if occupied.intersection(set(slots_to_use)):
                    continue

            room = None
            if not is_elective:
                if code in course_room_map:
                    room = course_room_map[code]
                else:
                    if session_type == "P":
                        if not labs:
                            print(f"No labs available for {code}")
                            return False
                        room = random.choice(labs)
                    else:
                        if not classrooms:
                            print(f"No classrooms available for {code}")
                            return False
                        room = random.choice(classrooms)
                    course_room_map[code] = room

            for i, s in enumerate(slots_to_use):
                if session_type == "L":
                    timetable.at[day, s] = f"{code} ({room})" if (room and not is_elective) else code
                elif session_type == "T":
                    timetable.at[day, s] = f"{code}T ({room})" if (room and not is_elective) else f"{code}T"
                elif session_type == "P":
                    timetable.at[day, s] = f"{code} (Lab-{room})" if (room and not is_elective) else code
                else:
                    timetable.at[day, s] = f"{code} ({room})" if (room and not is_elective) else code

            if faculty:
                lecturer_busy[day].setdefault(faculty, set()).update(slots_to_use)
            if session_type == "P":
                labs_on_days.add(day)
            return True
    return False

def merge_and_style_cells(filename):
    wb = load_workbook(filename)
    ws = wb.active

    course_colors = {}
    available_colors = colors.copy()
    random.shuffle(available_colors)

    slot_col_start = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        col = slot_col_start
        while col <= max_col:
            raw_val = ws.cell(row=row_idx, column=col).value
            if raw_val is None or str(raw_val).strip() == "":
                col += 1
                continue

            cell_value = str(raw_val).strip()
            merge_cols = [col]

            if "(" in cell_value:
                if "Lab" in cell_value:
                    expected_dur = 2.0
                elif cell_value.endswith("T") or "T " in cell_value or "T(" in cell_value:
                    expected_dur = 1.0
                else:
                    expected_dur = 1.5
            else:
                expected_dur = 1.5

            slot_index = (col - slot_col_start)
            total_dur = slot_durations[slot_keys[slot_index]] if 0 <= slot_index < len(slot_keys) else 0.0

            next_col = col + 1
            while next_col <= max_col:
                next_raw = ws.cell(row=row_idx, column=next_col).value
                next_val = str(next_raw).strip() if next_raw is not None else ""
                if next_val == cell_value:
                    sn_idx = (next_col - slot_col_start)
                    if 0 <= sn_idx < len(slot_keys):
                        total_dur += slot_durations[slot_keys[sn_idx]]
                    merge_cols.append(next_col)
                    if total_dur + 1e-9 >= expected_dur:
                        break
                    next_col += 1
                else:
                    break

            if len(merge_cols) > 1:
                ws.merge_cells(start_row=row_idx, start_column=merge_cols[0],
                               end_row=row_idx, end_column=merge_cols[-1])

            cell = ws.cell(row=row_idx, column=merge_cols[0])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)

            raw_course_name = cell_value.split()[0] if cell_value.split() else cell_value
            raw_course_name = raw_course_name.replace("T", "").replace("(", "").strip()

            if raw_course_name not in course_colors:
                if available_colors:
                    course_colors[raw_course_name] = available_colors.pop()
                else:
                    r = lambda: random.randint(150, 255)
                    course_colors[raw_course_name] = f"{r():02X}{r():02X}{r():02X}"
            fill_color = course_colors[raw_course_name]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for c in merge_cols:
                ws.cell(row=row_idx, column=c).border = thin_border

            col = merge_cols[-1] + 1

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(filename)

# ================== MAIN TIMETABLE FUNCTION ==================
def generate_timetable(courses_to_allocate, filename):
    timetable = pd.DataFrame("", index=days, columns=slot_keys)
    lecturer_busy = {day: {} for day in days}
    course_room_map = {}
    labs_on_days = set()
    unscheduled = []  # track unscheduled

    electives = [c for c in courses_to_allocate if safe_str(c.get("Elective", "")) == "1"]
    non_electives = [c for c in courses_to_allocate if safe_str(c.get("Elective", "")) != "1"]

    if electives:
        chosen = random.choice(electives)
        elective_course = {
            "Course_Code": "Elective",
            "Faculty": chosen.get("Faculty", ""),
            "L-T-P-S-C": chosen.get("L-T-P-S-C", "0-0-0-0-0"),
            "Elective": "1"
        }
        non_electives.append(elective_course)

    for course in non_electives:
        faculty = safe_str(course.get("Faculty", ""))
        code = safe_str(course.get("Course_Code", "UNKNOWN"))
        is_elective = (code == "Elective")
        L, T, P, S, C = parse_ltp(course.get("L-T-P-S-C", "0-0-0-0-0"))

        # ---- Lectures ----
        lecture_hours_remaining = L
        attempts = 0
        while lecture_hours_remaining > 1e-9 and attempts < MAX_ATTEMPTS:
            attempts += 1
            allocated = False
            alloc_target = 1.5 if lecture_hours_remaining >= 1.5 else 1.0
            for day in days:
                if lecture_hours_remaining <= 1e-9:
                    break
                if allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code, alloc_target, "L", is_elective, labs_on_days):
                    lecture_hours_remaining -= alloc_target
                    allocated = True
                    break
            if not allocated and alloc_target == 1.5:
                for day in days:
                    if allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code, 1.0, "L", is_elective, labs_on_days):
                        lecture_hours_remaining -= 1.0
                        allocated = True
                        break
        if lecture_hours_remaining > 1e-9:
            unscheduled.append({"Course_Code": code, "Faculty": faculty, "Type": "Lecture", "Unscheduled_Hours": lecture_hours_remaining})

        # ---- Tutorials ----
        tutorial_hours_remaining = T
        attempts = 0
        while tutorial_hours_remaining > 1e-9 and attempts < MAX_ATTEMPTS:
            attempts += 1
            allocated = False
            for day in days:
                if allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code, 1.0, "T", is_elective, labs_on_days):
                    tutorial_hours_remaining -= 1.0
                    allocated = True
                    break
        if tutorial_hours_remaining > 1e-9:
            unscheduled.append({"Course_Code": code, "Faculty": faculty, "Type": "Tutorial", "Unscheduled_Hours": tutorial_hours_remaining})

        # ---- Practicals ----
        practical_hours_remaining = P
        attempts = 0
        while practical_hours_remaining > 1e-9 and attempts < MAX_ATTEMPTS:
            attempts += 1
            allocated = False
            if practical_hours_remaining >= 2.0:
                for day in days:
                    if allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code, 2.0, "P", is_elective, labs_on_days):
                        practical_hours_remaining -= 2.0
                        allocated = True
                        break
            if not allocated:
                for day in days:
                    if allocate_session(timetable, lecturer_busy, course_room_map, day, faculty, code, 1.0, "P", is_elective, labs_on_days):
                        practical_hours_remaining -= 1.0
                        allocated = True
                        break
        if practical_hours_remaining > 1e-9:
            unscheduled.append({"Course_Code": code, "Faculty": faculty, "Type": "Practical", "Unscheduled_Hours": practical_hours_remaining})

    # Clean excluded slots
    for day in days:
        for slot in excluded_slots:
            if slot in timetable.columns:
                timetable.at[day, slot] = ""

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outname = f"{timestamp}_{filename}"
    timetable.to_excel(outname, index=True)
    merge_and_style_cells(outname)

    # ---- Save unscheduled ----
    if unscheduled:
        df_unscheduled = pd.DataFrame(unscheduled)
        unsched_filename = f"{timestamp}_unscheduled_courses.xlsx"
        df_unscheduled.to_excel(unsched_filename, index=False)
        print(f"⚠️ Some sessions could not be scheduled. Saved in {unsched_filename}")

    # ---- Append course info ----
    wb = load_workbook(outname)
    ws = wb.active
    start_row = ws.max_row + 3

    friendly_headers = {
        "Course_Code": "Course Code",
        "Course_Title": "Course Title",
        "L-T-P-S-C": "L-T-P-S-C",
        "Faculty": "Faculty",
        "Elective": "Elective"
    }

    headers = [friendly_headers.get(k, k) for k in courses_to_allocate[0].keys() if k != "Semester_Half"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for r_idx, course in enumerate(courses_to_allocate, start=start_row + 1):
        col_idx = 1
        for key in course.keys():
            if key == "Semester_Half":
                continue
            val = course[key]
            if key == "Elective":
                val = "Yes" if str(val).strip() == "1" else "No"
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    wb.save(outname)
    print(f"✅ Saved styled timetable with course info in {outname}")

# ================== SPLIT AND GENERATE ==================
def split_by_half(courses_list):
    first = [c for c in courses_list if safe_str(c.get("Semester_Half", "")) in ["1", "0"]]
    second = [c for c in courses_list if safe_str(c.get("Semester_Half", "")) in ["2", "0"]]
    return first, second

c1_first, c1_second = split_by_half(coursesA)
c2_first, c2_second = split_by_half(coursesB)
c3_first, c3_second = split_by_half(coursesECE)
c4_first, c4_second = split_by_half(coursesDSAI)

generate_timetable(c1_first, "timetable_first_halfCSEA.xlsx")
generate_timetable(c1_second, "timetable_second_halfCSEA.xlsx")
generate_timetable(c2_first, "timetable_first_halfCSEB.xlsx")
generate_timetable(c2_second, "timetable_second_halfCSEB.xlsx")
generate_timetable(c3_first, "timetable_first_halfECE.xlsx")
generate_timetable(c3_second, "timetable_second_halfECE.xlsx")
generate_timetable(c4_first, "timetable_first_halfDSAI.xlsx")
generate_timetable(c4_second, "timetable_second_halfDSAI.xlsx")


