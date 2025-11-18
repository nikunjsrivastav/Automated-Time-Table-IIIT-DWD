
import pandas as pd
import json
import random
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

random.seed(12345)
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded = ["07:30-09:00", "10:30-10:45", "13:15-14:00", "15:30-15:40"]
colors = [
    "FFB3BA","BAE1FF","BAFFC9","FFFFBA","FFD8BA","E3BAFF","D0BAFF","FFCBA4",
    "C7FFD8","B8E1FF","F7FFBA","FFDFBA","E9BAFF","BAFFD9","FFE1BA","BAFFF2",
    "D1FFBA","B2D8F7","F2C2FF","C2FFD8","FFB8E1","D8FFB8","FFE3BA","BAE7FF",
    "E8BAFF","BAFFD6","FFF2BA","DAD7FF","BFFFE1","FFDAB8","E2FFBA","BAF7FF"
]
thin = Border(left=Side(style='thin'),
              right=Side(style='thin'),
              top=Side(style='thin'),
              bottom=Side(style='thin'))

with open("data/time_slots.json") as f:
    slots = json.load(f)["time_slots"]

def t2m(t):
    h, m = map(int, t.split(":"))
    return h*60 + m

slots_norm = [
    {
        "key": f"{s['start']}-{s['end']}",
        "start": s['start'],
        "end": s['end'],
        "dur": (t2m(s["end"]) - t2m(s["start"])) / 60.0
    }
    for s in slots
]

slots_norm.sort(key=lambda x: t2m(x["start"]))

slot_keys = [s["key"] for s in slots_norm]
slot_dur = {s["key"]: s["dur"] for s in slots_norm}

coursesAI = pd.read_csv("data/coursesCSEA-I.csv").to_dict(orient="records")
coursesBI = pd.read_csv("data/coursesCSEB-I.csv").to_dict(orient="records")
coursesA  = pd.read_csv("data/coursesCSEA-III.csv").to_dict(orient="records")
coursesB  = pd.read_csv("data/coursesCSEB-III.csv").to_dict(orient="records")
coursesV  = pd.read_csv("data/coursesCSE-V.csv").to_dict(orient="records")
coursesDSAI = pd.read_csv("data/coursesDSAI-III.csv").to_dict(orient="records")
coursesECE  = pd.read_csv("data/coursesECE-III.csv").to_dict(orient="records")
coursesVII  = pd.read_csv("data/courses7.csv").to_dict(orient="records")
coursesDSAI_I = pd.read_csv("data/coursesDSAI-I.csv").to_dict(orient="records")
coursesDSAI_V = pd.read_csv("data/coursesDSAI-V.csv").to_dict(orient="records")
coursesECE_I  = pd.read_csv("data/coursesECE-I.csv").to_dict(orient="records")
coursesECE_V  = pd.read_csv("data/coursesECE-V.csv").to_dict(orient="records")
rooms = pd.read_csv("data/rooms.csv")
rooms["Room_ID"] = rooms["Room_ID"].astype(str).str.strip() 
cls = rooms[rooms["Room_ID"].str.startswith('C')]
labs = rooms[
    (rooms["Room_ID"].str.startswith('L'))
]
try:
    reg = pd.read_csv("registrations.csv")
    reg.set_index("Course_Code", inplace=True)
except Exception:
    reg = None


def regd(c):
    try:
        return int(reg.at[c, "Registered"])
    except Exception:
        return 0

def s(v):
    if v is None: return ""
    if isinstance(v, float) and pd.isna(v): return ""
    return str(v).strip()

def ltp(sv):
    try:
        p = [x.strip() for x in sv.split("-")]
    except Exception:
        return [0,0,0,0,0]
    while len(p) < 5:
        p.append("0")
    return list(map(int, p[:5]))


pat = re.compile(r"^[A-Z]{1,5}\d{0,3}([+/\\-][A-Z]{1,5}\d{0,3})*$", re.I)

def valid(c):
    codes, err = [], []
    for x in c:
        code = s(x.get("Course_Code", ""))
        if not code:
            continue
        if code.upper() in {"NEW", "ELECTIVE"}:
            codes.append(code.upper())
            continue
        if not pat.match(code):
            err.append(code)
        codes.append(code.upper())
    dup = {x for x in codes if codes.count(x) > 1 and x not in {"NEW", "ELECTIVE"}}
    if dup:
        err += list(dup)
    return err
def room(code, n, lab=False):
    df = labs if lab else cls
    if df.empty:
        return None
    if n <= 0:
        return random.choice(df["Room_ID"].tolist())
    
    fdf = df[df["Capacity"].astype(int) >= n]
    
    if not fdf.empty:
        return fdf.sort_values("Capacity").iloc[0]["Room_ID"]

    return None

def free(tt, d, ex=False):
    fb, b = [], []
    for s_ in slot_keys:
        if not ex and s_ in excluded:
            if b:
                fb.append(b)
                b = []
            continue
        if tt.at[d, s_] == "":
            b.append(s_)
        else:
            if b:
                fb.append(b)
                b = []
    if b:
        fb.append(b)
    return fb

def alloc_specific(tt, busy, rm, day, slots_to_use, f, code, typ, elec, labsd, course_usage):
    # Validate slots and free-ness
    for s_ in slots_to_use:
        if s_ not in slot_keys or tt.at[day, s_] != "":
            return False

    # Enforce course usage rule
    if code not in course_usage[day]:
        course_usage[day][code] = {"L": 0, "T": 0, "P": 0}

    if course_usage[day][code][typ] >= 1:
        return False

    # Room allocation for non-elective
    r = None
    if not elec:
        key = (code, typ) 
        if key in rm: 
            r = rm[key] 
        else:
            r = room(code, regd(code), typ == "P")
            rm[key] = r
        if r is None:
            return False

    for s_ in slots_to_use:
        if r and not elec:
            if typ == "T":
                v = f"{code}T ({r})"
            elif typ == "P":
                v = f"{code} (Lab-{r})"
            else:
                v = f"{code} ({r})"
        else:
            if typ == "T":
                v = f"{code}T"
            else:
                v = code
        tt.at[day, s_] = v
        
    if f:
        busy[day].setdefault(f, set()).update(slots_to_use)
    if typ == "P":
        labsd.add(day)

    # Update course usage for the day
    course_usage[day][code][typ] += 1
    return True


def alloc(tt, busy, rm, d, f, code, h, typ="L", elec=False, labsd=set(), ex=False, preferred_slots=None, course_usage=None):
    """
    Alloc with per-course-per-day rule: accepts course_usage dict (created in generate)
    and ensures at most one L/T/P per course per day.
    """
    if course_usage is None:
        # safety: if not provided, create a local structure (but generate should always pass it)
        course_usage = {dd: {} for dd in days}

    # === RULE ENFORCEMENT: only one L/T/P per course per day ===
    if code not in course_usage[d]:
        course_usage[d][code] = {"L": 0, "T": 0, "P": 0}

    if course_usage[d][code][typ] >= 1:
        return False

    # Preferred slot handling (try exact requested slots first)
    if preferred_slots:
        pref_day, pref_slots = preferred_slots
        if pref_day == d:
            total = sum(slot_dur[s] for s in pref_slots)
            if total + 1e-9 >= h:
                if alloc_specific(tt, busy, rm, pref_day, pref_slots, f, code, typ, elec, labsd, course_usage):
                    return True
                        
    for blk in free(tt, d, ex):
        if sum(slot_dur[s] for s in blk) + 1e-9 < h:
            continue
        use = []
        dur = 0.0
        for s_ in blk:
            use.append(s_)
            dur += slot_dur[s_]
            if dur + 1e-9 >= h:
                break
        if not ex and any(s_ in excluded for s_ in use):
            continue
        if f and f in busy[d] and set(use) & busy[d][f]:
            continue
            
        r = None
        if not elec:
            key = (code, typ)
            if key in rm: 
                r = rm[key]
            else:
                r = room(code, regd(code), typ == "P")
                rm[key] = r

            if r is None:
                continue 

        for s_ in use:
            if r and not elec:
                if typ == "T":
                    v = f"{code}T ({r})"
                elif typ == "P":
                    v = f"{code} (Lab-{r})"
                else:
                    v = f"{code} ({r})"
            else:
                if typ == "T":
                    v = f"{code}T"
                else:
                    v = code
            tt.at[d, s_] = v
            
        if f:
            busy[d].setdefault(f, set()).update(use)
        if typ == "P":
            labsd.add(d)

        # === UPDATE COURSE DAILY USAGE ===
        course_usage[d][code][typ] += 1  

        return True
    return False
color_avail = colors.copy()
random.shuffle(color_avail)
color_map = {}

def get_color_for_course(course_code):
    k = course_code.strip().upper()
    if k == "":
        return None
    if k not in color_map:
        if color_avail:
            color_map[k] = color_avail.pop()
        else:
            color_map[k] = "CCCCCC"
    return color_map[k]

def merge_and_color(ws, courses):
    sc = 2
    mc = ws.max_column
    mr = ws.max_row
    valid_course_codes = {
        s(x.get("Course_Code", "")).replace("T", "").strip().upper()
        for x in courses if s(x.get("Course_Code", ""))
    }
    valid_course_codes |= {f"ELECTIVE{i}" for i in range(1, 60)}

    for col in range(2, mc + 1):
        cell = ws.cell(2, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for r in range(3, mr + 1):
        c = sc
        while c <= mc:
            raw = ws.cell(r, c).value
            if raw is None or str(raw).strip() == "":
                ws.cell(r, c).border = thin
                c += 1
                continue

            val = str(raw).strip()
            merge_cols = [c]

            if "(" in val:
                if "Lab" in val:
                    expected = 2.0
                elif val.endswith("T") or "T " in val or "T(" in val:
                    expected = 1.0
                else:
                    expected = 1.5
            else:
                expected = 1.5

            slot_index = c - sc
            total = 0.0
            if 0 <= slot_index < len(slot_keys):
                total = slot_dur[slot_keys[slot_index]]
            next_col = c + 1
            while next_col <= mc:
                next_raw = ws.cell(r, next_col).value
                next_val = str(next_raw).strip() if next_raw is not None else ""
                if next_val == val:
                    sn_idx = next_col - sc
                    if 0 <= sn_idx < len(slot_keys):
                        total += slot_dur[slot_keys[sn_idx]]
                    merge_cols.append(next_col)
                    if total + 1e-9 >= expected:
                        break
                    next_col += 1
                else:
                    break

            if len(merge_cols) > 1:
                ws.merge_cells(start_row=r, start_column=merge_cols[0], end_row=r, end_column=merge_cols[-1])

            cell = ws.cell(r, merge_cols[0])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)

            raw_course_name = val.split()[0] if val.split() else val
            raw_course_name = raw_course_name.replace("T", "").replace("(", "").strip().upper()

            if raw_course_name in valid_course_codes or raw_course_name.startswith("ELECTIVE"):
                fill_color = get_color_for_course(raw_course_name)
            else:
                fill_color = None

            for cc_ in merge_cols:
                cell_ref = ws.cell(r, cc_)
                cell_ref.border = thin
                cell_ref.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell_ref.font = Font(bold=True)
                if fill_color:
                    cell_ref.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            c = merge_cols[-1] + 1

    for col in ws.columns:
        maxl = 0
        cl = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            maxl = max(maxl, len(str(v)))
        ws.column_dimensions[cl].width = min(maxl + 2 if maxl else 8, 60)
def add_csv_legend_block(ws, csv_path, legend_title):
    ws.append([""])
    ws.append([""])
    ws.append([f"Legend - {legend_title}"])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    df = pd.read_csv(csv_path)
    expect_cols = ["Course_Code", "Course_Title", "L-T-P-S-C", "Faculty", "Semester_Half", "Elective"]
    for ec in expect_cols:
        if ec not in df.columns:
            alt = None
            low = ec.lower()
            for c in df.columns:
                if c.lower() == low:
                    alt = c
                    break
            if alt:
                df.rename(columns={alt: ec}, inplace=True)
            else:
                if ec == "Semester_Half":
                    df[ec] = 0
                elif ec == "Elective":
                    df[ec] = 0
                else:
                    df[ec] = ""

    df = df[expect_cols].copy()

    def map_sem(x):
        try:
            xi = int(x)
        except Exception:
            xi = 0
        if xi == 1: return "First Half"
        if xi == 2: return "Second Half"
        return "Full Sem"

    def map_elec(x):
        try:
            xi = int(x)
        except Exception:
            xi = 0
        return "Yes" if xi == 1 else "No"

    df["Semester_Half"] = df["Semester_Half"].apply(map_sem)
    df["Elective"] = df["Elective"].apply(map_elec)

    headers = ["Course Code", "Course Title", "L-T-P-S-C", "Faculty", "Semester Half", "Elective"]
    ws.append(headers)
    header_row = ws.max_row
    for i, _h in enumerate(headers, start=1):
        c = ws.cell(header_row, i)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
        c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, row in df.iterrows():
        rowvals = [s(row["Course_Code"]), s(row["Course_Title"]), s(row["L-T-P-S-C"]), s(row["Faculty"]), s(row["Semester_Half"]), s(row["Elective"])]
        ws.append(rowvals)
        for i in range(1, 7):
            cc = ws.cell(ws.max_row, i)
            cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc.border = thin

    ws.append([""])


def generate(courses, ws, label, seed, elective_sync):
    if valid(courses):
        return []

    ws.append([""])
    ws.append([label])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    tt = pd.DataFrame("", index=days, columns=slot_keys)
    busy = {d: {} for d in days}
    rm = {}
    labsd = set()

    # --------------------- NEW: course_usage dict ---------------------
    # Tracks per-day usage for each course: {"Monday": {"CS101": {"L":0,"T":0,"P":0}, ...}, ...}
    course_usage = {d: {} for d in days}
    # -----------------------------------------------------------------

    elec = [x for x in courses if s(x.get("Elective", "")) == "1"]
    non = [x for x in courses if s(x.get("Elective", "")) != "1"]

    baskets = {}
    elec_no_baskets = []
    for e in elec:
        b = s(e.get("ElectiveBasket", "0"))
        if b and b != "0":
            baskets.setdefault(b, []).append(e)
        else:
            elec_no_baskets.append(e)
    basket_reps = []
    for b, group in sorted(baskets.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        chosen = group[0]
        sync_identifier = f"BASKET_{b}"
        basket_reps.append({
            "Course_Code": f"Elective{b}",
            "Course_Title": chosen.get("Course_Title", "") or chosen.get("Course_Code", ""),
            "Faculty": chosen.get("Faculty", ""),
            "L-T-P-S-C": chosen.get("L-T-P-S-C", "0-0-0-0-0"),
            "Elective": "1",
            "ElectiveBasket": b,
            "_sync_name": sync_identifier
        })

    for e in elec_no_baskets:
        sync_n = s(e.get("Course_Title", "")) or s(e.get("Course_Code", ""))
        e["_sync_name"] = sync_n if sync_n else None

    elec_final = elec_no_baskets + basket_reps

    def place_course_list(course_list, start_idx_ref):
        placed_list = []
        for c in course_list:
            f = s(c.get("Faculty", ""))
            code = s(c.get("Course_Code", "UNKNOWN"))
            ele = (code.startswith("Elective") or s(c.get("Elective", "")) == "1")
            L, T, P, S, C = ltp(c.get("L-T-P-S-C", "0-0-0-0-0"))

            for h, typ in [(L, "L"), (T, "T"), (P, "P")]:
                attempts = 0
                while h > 1e-9 and attempts < 300:
                    if typ == "P":
                        a = 2.0 if h >= 2 else (1.5 if h >= 1.5 else 1.0)
                    else:
                        a = 1.5 if h >= 1.5 else 1.0

                    placed = False

                    sync_name = c.get("_sync_name", None)
                    if sync_name and sync_name in elective_sync:
                        pref = elective_sync[sync_name]
                        if alloc(tt, busy, rm, pref["day"], f, code, a, typ, ele, labsd, False, preferred_slots=(pref["day"], pref["slots"]), course_usage=course_usage):
                            h -= a
                            placed = True

                    if not placed:
                        if ele:
                            d_order = days[:]
                        else:
                            start_idx = start_idx_ref[0]
                            d_order = days[start_idx:] + days[:start_idx]
                            start_idx_ref[0] = (start_idx_ref[0] + 1) % len(days)

                        for d in d_order:
                            if alloc(tt, busy, rm, d, f, code, a, typ, ele, labsd, False, course_usage=course_usage):
                                h -= a
                                placed = True
                                break

                    if not placed:
                        for d in days:
                            if alloc(tt, busy, rm, d, f, code, a, typ, ele, labsd, True, course_usage=course_usage):
                                h -= a
                                placed = True
                                break

                    if placed and sync_name and sync_name not in elective_sync:
                        for dcheck in days:
                            slots_used = [s_ for s_ in slot_keys if tt.at[dcheck, s_].startswith(code)]
                            if slots_used:
                                accum = []
                                acc_dur = 0.0
                                for s_ in slots_used:
                                    accum.append(s_)
                                    acc_dur += slot_dur[s_]
                                    if acc_dur + 1e-9 >= a:
                                        elective_sync[sync_name] = {"day": dcheck, "slots": accum.copy()}
                                        break
                                if sync_name in elective_sync:
                                    break
                    attempts += 1
            placed_list.append(c)
        return placed_list

    start_idx_ref = [seed % len(days)]

    elec_placed = place_course_list(elec_final, start_idx_ref)
    non_placed = place_course_list(non, start_idx_ref)

    combined = non_placed + elec_placed

    ws.append(["Day"] + slot_keys)
    for d in days:
        ws.append([d] + [tt.at[d, s] for s in slot_keys])
    ws.append([""])

    return combined
def split(c):
    f = [x for x in c if s(x.get("Semester_Half", "")) in ["1", "0"]]
    s2 = [x for x in c if s(x.get("Semester_Half", "")) in ["2", "0"]]
    return f, s2

if __name__ == "__main__":

    wb = Workbook()
    seed = random.randint(0, 999999)

    ws1 = wb.active
    ws1.title = "CSE-I Timetable"
    sync_I_sem = {}

    cAf, cAs = split(coursesAI)
    cBf, cBs = split(coursesBI)

    csea_block = generate(cAf, ws1, "CSEA I First Half", seed + 0, sync_I_sem)
    csea_block2 = generate(cAs, ws1, "CSEA I Second Half", seed + 1, sync_I_sem)
    add_csv_legend_block(ws1, "data/coursesCSEA-I.csv", "CSEA I")

    cseb_block = generate(cBf, ws1, "CSEB I First Half", seed + 2, sync_I_sem)
    cseb_block2 = generate(cBs, ws1, "CSEB I Second Half", seed + 3, sync_I_sem)
    add_csv_legend_block(ws1, "data/coursesCSEB-I.csv", "CSEB I")

    combined_i_courses = (csea_block or []) + (csea_block2 or []) + (cseb_block or []) + (cseb_block2 or [])
    merge_and_color(ws1, combined_i_courses)

    sync_III_sem = {}

    ws2 = wb.create_sheet("CSE-III Timetable")
    c1f, c1s = split(coursesA)
    c2f, c2s = split(coursesB)

    csea3_block1 = generate(c1f, ws2, "CSEA III First Half", seed + 4, sync_III_sem)
    csea3_block2 = generate(c1s, ws2, "CSEA III Second Half", seed + 5, sync_III_sem)
    add_csv_legend_block(ws2, "data/coursesCSEA-III.csv", "CSEA III")

    cseb3_block1 = generate(c2f, ws2, "CSEB III First Half", seed + 6, sync_III_sem)
    cseb3_block2 = generate(c2s, ws2, "CSEB III Second Half", seed + 7, sync_III_sem)
    add_csv_legend_block(ws2, "data/coursesCSEB-III.csv", "CSEB III")

    combined_iii_courses = (csea3_block1 or []) + (csea3_block2 or []) + (cseb3_block1 or []) + (cseb3_block2 or [])
    merge_and_color(ws2, combined_iii_courses)

    ws4 = wb.create_sheet("DSAI-III Timetable")
    d1f, d1s = split(coursesDSAI)

    dsa_block1 = generate(d1f, ws4, "DSAI-III First Half", seed + 10, sync_III_sem)
    dsa_block2 = generate(d1s, ws4, "DSAI=III Second Half", seed + 11, sync_III_sem)
    add_csv_legend_block(ws4, "data/coursesDSAI-III.csv", "DSAI")

    combined_dsa_courses = (dsa_block1 or []) + (dsa_block2 or [])
    merge_and_color(ws4, combined_dsa_courses)

    ws5 = wb.create_sheet("ECE-III Timetable")
    e1f, e1s = split(coursesECE)

    ece_block1 = generate(e1f, ws5, "ECE-III First Half", seed + 12, sync_III_sem)
    ece_block2 = generate(e1s, ws5, "ECE-III Second Half", seed + 13, sync_III_sem)
    add_csv_legend_block(ws5, "data/coursesECE-III.csv", "ECE")

    combined_ece_courses = (ece_block1 or []) + (ece_block2 or [])
    merge_and_color(ws5, combined_ece_courses)

    sync_V_sem = {}

    ws3 = wb.create_sheet("CSE-V Timetable")
    c5f, c5s = split(coursesV)

    c5_block1 = generate(c5f, ws3, "CSE-V First Half", seed + 8, sync_V_sem)
    c5_block2 = generate(c5s, ws3, "CSE-V Second Half", seed + 9, sync_V_sem)
    add_csv_legend_block(ws3, "data/coursesCSE-V.csv", "CSE V")

    combined_v_courses = (c5_block1 or []) + (c5_block2 or [])
    merge_and_color(ws3, combined_v_courses)

    ws6 = wb.create_sheet("DSAI 7TH-SEM Timetable")
    s7f, s7s = split(coursesVII)

    s7_block1 = generate(s7f, ws6, "DSAI 7TH-SEM First Half", seed + 14, {})
    s7_block2 = generate(s7s, ws6, "DSAI 7TH-SEM Second Half", seed + 15, {})
    add_csv_legend_block(ws6, "data/courses7.csv", "7TH SEM")

    combined_7_courses = (s7_block1 or []) + (s7_block2 or [])
    merge_and_color(ws6, combined_7_courses)

    ws7 = wb.create_sheet("DSAI-I Timetable")
    d1f_i, d1s_i = split(coursesDSAI_I) 
    dsai1_block1 = generate(d1f_i, ws7, "DSAI-I First Half", seed + 16, {})
    dsai1_block2 = generate(d1s_i, ws7, "DSAI-I Second Half", seed + 17, {})
    add_csv_legend_block(ws7, "data/coursesDSAI-I.csv", "DSAI I")

    combined_dsai1_courses = (dsai1_block1 or []) + (dsai1_block2 or [])
    merge_and_color(ws7, combined_dsai1_courses)

    ws8 = wb.create_sheet("DSAI-V Timetable")
    d5f_v, d5s_v = split(coursesDSAI_V) 

    dsai5_block1 = generate(d5f_v, ws8, "DSAI-V First Half", seed + 18, sync_V_sem)
    dsai5_block2 = generate(d5s_v, ws8, "DSAI-V Second Half", seed + 19, sync_V_sem)
    add_csv_legend_block(ws8, "data/coursesDSAI-V.csv", "DSAI V")

    combined_dsai5_courses = (dsai5_block1 or []) + (dsai5_block2 or [])
    merge_and_color(ws8, combined_dsai5_courses)

    ws9 = wb.create_sheet("ECE-I Timetable")
    e1f_i, e1s_i = split(coursesECE_I) 

    ece1_block1 = generate(e1f_i, ws9, "ECE-I First Half", seed + 20, {})
    ece1_block2 = generate(e1s_i, ws9, "ECE-I Second Half", seed + 21, {})
    add_csv_legend_block(ws9, "data/coursesECE-I.csv", "ECE I")

    combined_ece1_courses = (ece1_block1 or []) + (ece1_block2 or [])
    merge_and_color(ws9, combined_ece1_courses)

    ws10 = wb.create_sheet("ECE-V Timetable")
    e5f_v, e5s_v = split(coursesECE_V) 

    ece5_block1 = generate(e5f_v, ws10, "ECE-V First Half", seed + 22, sync_V_sem)
    ece5_block2 = generate(e5s_v, ws10, "ECE-V Second Half", seed + 23, sync_V_sem)
    add_csv_legend_block(ws10, "data/coursesECE-V.csv", "ECE V")

    combined_ece5_courses = (ece_block1 or []) + (ece_block2 or [])
    merge_and_color(ws10, combined_ece5_courses)
    
    name = f"Balanced_Timetable_latest.xlsx"
    wb.save(name)
    print("✅ Evenly balanced timetable saved in", name)
