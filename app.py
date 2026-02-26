import streamlit as st
import streamlit.components.v1 as components # <-- ADD THIS
import pandas as pd
import json
import random
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==========================================
# 1. UI SETUP
# ==========================================
st.set_page_config(page_title="College Timetable Generator", layout="wide")

# ==========================================
# 2. GLOBALS & CONSTANTS
# ==========================================
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded = ["07:30-09:00", "10:30-10:45", "13:15-14:00", "17:30-18:30"]

colors = [
    "FFB3BA","BAE1FF","BAFFC9","FFFFBA","FFD8BA","E3BAFF","D0BAFF","FFCBA4",
    "C7FFD8","B8E1FF","F7FFBA","FFDFBA","E9BAFF","BAFFD9","FFE1BA","BAFFF2",
    "D1FFBA","B2D8F7","F2C2FF","C2FFD8","FFB8E1","D8FFB8","FFE3BA","BAE7FF",
    "E8BAFF","BAFFD6","FFF2BA","DAD7FF","BFFFE1","FFDAB8","E2FFBA","BAF7FF"
]

thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
lab_prefix_for_class_prefix = {"C1": "L1", "C2": "L2", "C3": "L3", "C4": "L4"}

# Dynamic globals defined during generation
slot_keys = []
slot_dur = {}
cls = pd.DataFrame()
labs = pd.DataFrame()
reg = None
color_avail = []
color_map = {}

# ==========================================
# 3. CORE ALGORITHM FUNCTIONS
# ==========================================
def t2m(t):
    h, m = map(int, t.split(":"))
    return h*60 + m

def regd(c):
    try:
        return int(reg.at[c, "Registered"])
    except Exception:
        return 0

def s(v):
    if v is None: return ""
    if isinstance(v, float) and pd.isna(v): return ""
    return str(v).strip()

def ltp(sv):
    try:
        p = [x.strip() for x in sv.split("-")]
    except Exception:
        return [0,0,0,0,0]
    while len(p) < 5:
        p.append("0")
    return list(map(int, p[:5]))

pat = re.compile(r"^[A-Z]{1,5}\d{0,3}([+/\\-][A-Z]{1,5}\d{0,3})*$", re.I)
def valid(c):
    codes_all, codes_core, err = [], [], []
    for x in c:
        code = s(x.get("Course_Code", ""))
        if not code: continue
        upper = code.upper()
        if upper in {"NEW", "ELECTIVE"}:
            codes_all.append(upper)
            continue
        if not pat.match(code):
            err.append(code)
        codes_all.append(upper)
        elec_flag = s(x.get("Elective", "")) == "1"
        if not elec_flag:
            codes_core.append(upper)
    dup_core = {x for x in codes_core if codes_core.count(x) > 1}
    if dup_core:
        err += list(dup_core)
    return err

def is_combined_course(code, rm):
    return (code, "L") in rm and rm[(code, "L")] == "C004"

def room_candidates(lab=False, prefix=None, lab_prefix=None):
    df = labs if lab else cls
    if df.empty: return []
    cand = df.copy()
    if prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(prefix.upper())]
        if not c.empty: cand = c
        else: cand = df.copy()
    if lab and lab_prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(lab_prefix.upper())]
        if not c.empty: cand = c
    return cand["Room_ID"].tolist()

def pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=None, rr_state=None):
    if not candidates: return None
    ordered = candidates
    if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
        idx = rr_state.get(rr_state_key, 0) % len(candidates)
        ordered = candidates[idx:] + candidates[:idx]
    for cand in ordered:
        used = room_busy.get(day, {}).get(cand, set())
        if not (set(slots_to_use) & used):
            if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
                rr_state[rr_state_key] = (rr_state.get(rr_state_key, 0) + 1) % len(candidates)
            return cand
    return None

def free(tt, d, ex=False):
    fb, b = [], []
    for s_ in slot_keys:
        if not ex and s_ in excluded:
            if b: fb.append(b); b = []
            continue
        if tt.at[d, s_] == "":
            b.append(s_)
        else:
            if b: fb.append(b); b = []
    if b: fb.append(b)
    return fb

def alloc_specific(tt, busy, rm, room_busy, day, slots_to_use, f, code, typ, elec, labsd, course_usage, class_prefix=None, rr_state=None, hide_c004=False):
    for s_ in slots_to_use:
        if s_ not in slot_keys or tt.at[day, s_] != "": return False
            
    if code not in course_usage[day]:
        course_usage[day][code] = {"L":0,"T":0,"P":0}

    usage = course_usage[day][code]
    if typ == "P" and elec:
        pass
    else:
        if typ == "P":
            if usage["P"] >= 1: return False
        else:
            if (usage["L"] + usage["T"]) >= 1: return False
            
    r = None
    if not elec:
        key = (code, typ)
        if key in rm:
            candidate = rm[key]
            if candidate != "C004": 
                used = room_busy.get(day, {}).get(candidate, set())
                if set(slots_to_use) & used: return False
            r = candidate
        else:
            if typ == "P" and elec: r = None
            elif typ == "P":
                lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref)
            else:
                candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None)
            r = pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=class_prefix, rr_state=rr_state)
            if r is None: return False
            rm[key] = r

    for s_ in slots_to_use:
        if is_combined_course(code, rm):
            if hide_c004:
                v = f"{code} (Lab)" if typ == "P" else (f"{code}T" if typ == "T" else f"{code}")
            else:
                v = f"{code} (Lab)" if typ == "P" else (f"{code}T (C004)" if typ == "T" else f"{code} (C004)")
        else:
            if r and not elec:
                v = f"{code}(Lab)" if (elec and typ == "P") else (f"{code}T ({r})" if typ == "T" else (f"{code} (Lab-{r})" if typ == "P" else f"{code} ({r})"))
            else:
                v = f"{code}(Lab)" if (elec and typ == "P") else (f"{code}T" if typ == "T" else code)
        tt.at[day, s_] = v

    if f: busy[day].setdefault(f, set()).update(slots_to_use)
    if r: room_busy.setdefault(day, {}).setdefault(r, set()).update(slots_to_use)
    if typ == "P": labsd.add(day)
    course_usage[day][code][typ] += 1
    return True

def alloc(tt, busy, rm, room_busy, d, f, code, h, typ="L", elec=False, labsd=set(), ex=False, preferred_slots=None, course_usage=None, class_prefix=None, rr_state=None, hide_c004=False):
    if course_usage is None: course_usage = {dd:{} for dd in days}
    if code not in course_usage[d]: course_usage[d][code] = {"L":0,"T":0,"P":0}
    usage = course_usage[d][code]

    if typ == "P":
        if usage["P"] >= 1: return False
    else:
        if (usage["L"] + usage["T"]) >= 1: return False

    if preferred_slots:
        pref_day, pref_slots = preferred_slots
        if pref_day == d:
            total = sum(slot_dur[s] for s in pref_slots)
            if total + 1e-9 >= h:
                if alloc_specific(tt, busy, rm, room_busy, pref_day, pref_slots, f, code, typ, elec, labsd, course_usage, class_prefix=class_prefix, rr_state=rr_state, hide_c004=hide_c004):
                    return True

    for blk in free(tt, d, ex):
        if sum(slot_dur[s] for s in blk) + 1e-9 < h: continue
        use = []; dur = 0.0
        for s_ in blk:
            use.append(s_); dur += slot_dur[s_]
            if dur + 1e-9 >= h: break
        if not ex and any(s_ in excluded for s_ in use): continue
        if f and f in busy[d] and (set(use) & busy[d][f]): continue

        if not elec:
            key = (code, typ)
            if key in rm:
                r = rm[key]
                if r != "C004":
                    used = room_busy.get(d, {}).get(r, set())
                    if set(use) & used: continue
            else:
                if typ == "P" and elec: r = None
                elif typ == "P":
                    lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                    candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=lab_pref, rr_state=rr_state)
                else:
                    candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=class_prefix, rr_state=rr_state)
                if r is None: continue
                rm[(code, typ)] = r
        else:
            r = None

        for s_ in use:
            if is_combined_course(code, rm):
                if hide_c004:
                    v = f"{code}(Lab)" if typ == "P" else (f"{code}T" if typ == "T" else f"{code}")
                else:
                    v = f"{code} (Lab)" if typ == "P" else (f"{code}T (C004)" if typ == "T" else f"{code} (C004)")
            else:
                if r and not elec:
                    v = f"{code}(Lab)" if (elec and typ == "P") else (f"{code}T ({r})" if typ == "T" else (f"{code} (Lab-{r})" if typ == "P" else f"{code} ({r})"))
                else:
                    v = f"{code}(Lab)" if (elec and typ == "P") else (f"{code}T" if typ == "T" else code)
            tt.at[d, s_] = v
        
        if f: busy[d].setdefault(f, set()).update(use)
        if r: room_busy.setdefault(d, {}).setdefault(r, set()).update(use)
        if typ == "P": labsd.add(d)
        course_usage[d][code][typ] += 1
        return True
    return False

def get_all_valid_free_slots(tt):
    valid = []
    for d in reversed(days):
        for s_ in reversed(slot_keys):
            if s_ in excluded: continue
            if tt.at[d, s_] == "": valid.append((d, s_))
    return valid

def get_all_excluded_free_slots(tt):
    exs = []
    for d in reversed(days):
        for s_ in reversed(slot_keys):
            if s_ not in excluded: continue
            if tt.at[d, s_] == "": exs.append((d, s_))
    return exs

def extract_contiguous_blocks(slot_list):
    blocks = []
    i = 0
    while i < len(slot_list):
        d0, s0 = slot_list[i]
        cur_day = d0
        cur_slots = [s0]
        i += 1
        while i < len(slot_list) and slot_list[i][0] == cur_day:
            cur_slots.append(slot_list[i][1]); i += 1
        blocks.append((cur_day, cur_slots))
    return blocks

def try_allocate_chunk_from_block(tt, busy, rm, room_busy, labsd, course_usage, code, faculty, typ, need, day, slots, class_prefix=None, rr_state=None, hide_c004=False):
    n = len(slots)
    for i in range(n):
        accum = 0.0; sub = []
        for j in range(i, n):
            sub.append(slots[j]); accum += slot_dur[slots[j]]
            if accum + 1e-9 >= need:
                for s_ in sub:
                    if tt.at[day, s_] != "": break
                else:
                    ok = alloc_specific(tt, busy, rm, room_busy, day, sub, faculty, code, typ, False, labsd, course_usage, class_prefix=class_prefix, rr_state=rr_state, hide_c004=hide_c004)
                    if ok:
                        new_slots = slots[:i] + slots[j+1:]
                        return new_slots
                break
    return None

def assign_combined_precise_durations(tt, busy, rm, room_busy, labsd, course_usage, combined_core, rr_state=None, hide_c004=False):
    if not combined_core: return []
    combined_list = []
    chunks_map = {}
    for c in combined_core:
        code = s(c.get("Course_Code", ""))
        if not code: continue
        rm[(code, "L")] = "C004"; rm[(code, "T")] = "C004"; rm[(code, "P")] = "C004"
        L, T, P, _, _ = ltp(c.get("L-T-P-S-C", "0-0-0-0-0"))
        ch = []
        rem = float(L)
        while rem > 1e-9:
            if rem >= 1.5: ch.append((1.5, "L")); rem -= 1.5
            else: ch.append((1.0, "L")); rem -= 1.0
        rem = float(T)
        while rem > 1e-9: ch.append((1.0, "T")); rem -= 1.0
        rem = float(P)
        while rem > 1e-9:
            if rem >= 2.0: ch.append((2.0, "P")); rem -= 2.0
            elif rem >= 1.5: ch.append((1.5, "P")); rem -= 1.5
            else: ch.append((1.0, "P")); rem -= 1.0
        chunks_map[code] = sorted(ch, key=lambda x: -x[0])
        combined_list.append((code, c))

    valid_slots = get_all_valid_free_slots(tt)
    valid_blocks = extract_contiguous_blocks(valid_slots)
    excluded_slots = get_all_excluded_free_slots(tt)
    excluded_blocks = extract_contiguous_blocks(excluded_slots)
    placed = []

    for code, c in combined_list:
        chunks = chunks_map[code]; faculty = s(c.get("Faculty", ""))
        days_used = set()
        for need, typ in chunks:
            allocated = False
            for idx, (day, slots) in enumerate(valid_blocks):
                if day in days_used: continue
                new_slots = try_allocate_chunk_from_block(tt, busy, rm, room_busy, labsd, course_usage, code, faculty, typ, need, day, slots, class_prefix="C0", rr_state=rr_state, hide_c004=hide_c004)
                if new_slots is not None:
                    valid_blocks[idx] = (day, new_slots); days_used.add(day); allocated = True; break
            if not allocated:
                for idx, (day, slots) in enumerate(excluded_blocks):
                    if day in days_used: continue
                    new_slots = try_allocate_chunk_from_block(tt, busy, rm, room_busy, labsd, course_usage, code, faculty, typ, need, day, slots, class_prefix="C0", rr_state=rr_state, hide_c004=hide_c004)
                    if new_slots is not None:
                        excluded_blocks[idx] = (day, new_slots); days_used.add(day); allocated = True; break
        placed.append(code)
    return placed

def get_color_for_course(course_code):
    k = course_code.strip().upper()
    if k == "": return None
    if k not in color_map:
        if color_avail: color_map[k] = color_avail.pop()
        else: color_map[k] = "CCCCCC"
    return color_map[k]

def merge_and_color(ws, courses):
    sc = 2; mc = ws.max_column; mr = ws.max_row
    valid_course_codes = {s(x.get("Course_Code","")).replace("T","").strip().upper() for x in courses if s(x.get("Course_Code",""))}
    valid_course_codes |= {f"ELECTIVE{i}" for i in range(1,60)}
    
    for r in range(1, mr+1):
        first_col_val = str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else ""
        
        # Format the "Day" / Timeslot headers
        if first_col_val == "Day":
            for col in range(1, mc+1):
                cell = ws.cell(r, col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin
            continue
        
        # Apply merge and color ONLY to timetable rows (Monday-Friday)
        if first_col_val in days:
            c = sc
            while c <= mc:
                raw = ws.cell(r, c).value
                if raw is None or str(raw).strip() == "":
                    ws.cell(r, c).border = thin; c += 1; continue
                
                val = str(raw).strip(); merge_cols = [c]
                if "(" in val:
                    if "Lab" in val: expected = 2.0
                    elif val.endswith("T") or "T " in val or "T(" in val: expected = 1.0
                    else: expected = 1.5
                else: expected = 1.5
                
                slot_index = c - sc; total = 0.0
                if 0 <= slot_index < len(slot_keys):
                    total = slot_dur[slot_keys[slot_index]]
                next_col = c + 1
                while next_col <= mc:
                    next_raw = ws.cell(r, next_col).value
                    next_val = str(next_raw).strip() if next_raw is not None else ""
                    if next_val == val:
                        sn_idx = next_col - sc
                        if 0 <= sn_idx < len(slot_keys):
                            total += slot_dur[slot_keys[sn_idx]]
                        merge_cols.append(next_col)
                        if total + 1e-9 >= expected: break
                        next_col += 1
                    else: break
                    
                if len(merge_cols) > 1:
                    ws.merge_cells(start_row=r, start_column=merge_cols[0], end_row=r, end_column=merge_cols[-1])
                    
                cell = ws.cell(r, merge_cols[0])
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(bold=True)
                
                raw_course_name = val.split()[0] if val.split() else val
                raw_course_name = raw_course_name.replace("T","").replace("(","").strip().upper()
                fill_color = get_color_for_course(raw_course_name) if (raw_course_name in valid_course_codes or raw_course_name.startswith("ELECTIVE")) else None
                
                for cc_ in merge_cols:
                    cell_ref = ws.cell(r, cc_)
                    cell_ref.border = thin
                    cell_ref.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell_ref.font = Font(bold=True)
                    if fill_color: cell_ref.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                c = merge_cols[-1] + 1
                
    for col in ws.columns:
        maxl = 0; cl = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None: continue
            maxl = max(maxl, len(str(v)))
        ws.column_dimensions[cl].width = min(maxl + 2 if maxl else 8, 60)
def add_csv_legend_block(ws, df, legend_title, room_prefix=None, elective_room_map=None):
    if elective_room_map is None: elective_room_map = {}
    if df.empty: return # Skip if no data loaded
    
    ws.append([""]); ws.append([""]); ws.append([f"Legend - {legend_title}"])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    expect_cols = ["Course_Code", "Course_Title", "L-T-P-S-C", "Faculty", "Semester_Half", "Elective", "ElectiveBasket"]
    for ec in expect_cols:
        if ec not in df.columns:
            alt = None; low = ec.lower()
            for c in df.columns:
                if c.lower() == low:
                    alt = c; break
            if alt: df.rename(columns={alt: ec}, inplace=True)
            else:
                if ec == "Semester_Half": df[ec] = 0
                elif ec == "Elective": df[ec] = 0
                else: df[ec] = ""

    df = df[["Course_Code", "Course_Title", "L-T-P-S-C", "Faculty", "Semester_Half", "Elective", "ElectiveBasket"]].copy()

    def map_sem(x):
        try: xi = int(x)
        except Exception: xi = 0
        if xi == 1: return "First Half"
        if xi == 2: return "Second Half"
        return "Full Sem"
    def map_elec(x):
        try: xi = int(x)
        except Exception: xi = 0
        return "Yes" if xi == 1 else "No"

    df["Semester_Half"] = df["Semester_Half"].apply(map_sem)
    df["Elective"] = df["Elective"].apply(map_elec)

    all_classrooms = cls["Room_ID"].tolist()
    master_pool = sorted(list(set(all_classrooms)))
    random.shuffle(master_pool)

    elective_rooms = []
    for _, row in df.iterrows():
        if row["Elective"] == "Yes":
            basket = str(row.get("ElectiveBasket", "")).strip()
            if basket and basket != "0":
                sync_name = f"{row['Course_Code']}_B{basket}"
            else:
                sync_name = row["Course_Code"]
            if sync_name in elective_room_map:
                chosen = elective_room_map[sync_name]
            else:
                taken_rooms = set(elective_room_map.values())
                candidates = [r for r in master_pool if r not in taken_rooms]
                chosen = candidates[0] if candidates else random.choice(master_pool)
                elective_room_map[sync_name] = chosen
            elective_rooms.append(f"{chosen}")
        else:
            elective_rooms.append("")

    df["Elective Room"] = elective_rooms
    headers = ["Course Code","Course Title","L-T-P-S-C","Faculty","Semester Half","Elective","Elective Basket","Elective Room"]
    ws.append(headers); header_row = ws.max_row
    for i, _h in enumerate(headers, start=1):
        c = ws.cell(header_row, i); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin; c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, row in df.iterrows():
        rowvals = [s(row["Course_Code"]), s(row["Course_Title"]), s(row["L-T-P-S-C"]), s(row["Faculty"]), s(row["Semester_Half"]), s(row["Elective"]), s(row["ElectiveBasket"]), row["Elective Room"]]
        ws.append(rowvals)
        for i in range(1, 9): # Fixed layout boundary to include all 8 columns
            cc = ws.cell(ws.max_row, i); cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cc.border = thin
    ws.append([""])

def generate(courses, ws, label, seed, elective_sync, room_prefix=None, elective_room_map=None, room_busy_global=None, hide_c004=False):
    if elective_room_map is None: elective_room_map = {}
    if valid(courses): return ([], [])
    
    ws.append([""]); ws.append([label])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    
    tt = pd.DataFrame("", index=days, columns=slot_keys)
    busy = {d:{} for d in days}
    room_busy = room_busy_global if room_busy_global is not None else {d:{} for d in days}
    rm = {}; labsd = set(); course_usage = {d:{} for d in days}; rr_state = {}

    elec = [x for x in courses if s(x.get("Elective","")) == "1"]
    combined_core = [x for x in courses if s(x.get("Elective","")) != "1" and s(x.get("Is_Combined","0")) == "1"]
    regular_core = [x for x in courses if s(x.get("Elective","")) != "1" and s(x.get("Is_Combined","0")) != "1"]

    baskets = {}; elec_no_baskets = []
    for e in elec:
        b = s(e.get("ElectiveBasket","0"))
        if b and b != "0": baskets.setdefault(b,[]).append(e)
        else: elec_no_baskets.append(e)
        
    basket_reps = []
    for b, group in sorted(baskets.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        chosen = group[0]
        basket_reps.append({
            "Course_Code": f"Elective{b}",
            "Course_Title": chosen.get("Course_Title","") or chosen.get("Course_Code",""),
            "Faculty": chosen.get("Faculty",""),
            "L-T-P-S-C": chosen.get("L-T-P-S-C","0-0-0-0-0"),
            "Elective": "1",
            "ElectiveBasket": b,
            "_sync_name": f"{chosen.get('Course_Code')}_B{b}"
        })

    for e in elec_no_baskets:
        basket = s(e.get("ElectiveBasket","0"))
        sync_n = f"{s(e.get('Course_Code'))}_B{basket}" if basket and basket != "0" else s(e.get("Course_Code"))
        e["_sync_name"] = sync_n if sync_n else None
        
    elec_final = elec_no_baskets + basket_reps
    for c in combined_core:
        code = s(c.get("Course_Code",""))
        rm[(code,"L")] = "C004"; rm[(code,"T")] = "C004"; rm[(code,"P")] = "C004"

    failed = []

    def place_course_list(course_list, start_idx_ref):
        placed_list = []
        for c in course_list:
            f = s(c.get("Faculty",""))
            code = s(c.get("Course_Code","UNKNOWN"))
            is_elec_flag = (code.startswith("Elective") or s(c.get("Elective","")) == "1")
            L, T, P, S, Cc = ltp(c.get("L-T-P-S-C","0-0-0-0-0"))
            for h, typ in [(L,"L"), (T,"T"), (P,"P")]:
                attempts = 0
                while h > 1e-9 and attempts < 400:
                    if typ == "P": a = 2.0 if h >= 2 else (1.5 if h >= 1.5 else 1.0)
                    else: a = 1.5 if h >= 1.5 else 1.0
                    placed = False
                    sync_name = c.get("_sync_name", None)

                    if is_elec_flag and sync_name and sync_name in elective_room_map:
                        for ttkey in [("L"), ("T"), ("P")]:
                            rm[(code, ttkey)] = elective_room_map[sync_name]

                    if sync_name and sync_name in elective_sync:
                        pref = elective_sync[sync_name]
                        if alloc(tt, busy, rm, room_busy, pref["day"], f, code, a, typ, is_elec_flag, labsd, False, preferred_slots=(pref["day"], pref["slots"]), course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state, hide_c004=hide_c004):
                            h -= a; placed = True

                    if not placed:
                        for i in range(5):
                            if is_elec_flag: d_order = days[:]
                            else:
                                start_idx = start_idx_ref[0]
                                d_order = days[start_idx:] + days[:start_idx]
                                start_idx_ref[0] = (start_idx_ref[0] + 1) % len(days)
                            for d in d_order:
                                if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, False, course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state, hide_c004=hide_c004):
                                    h -= a; placed = True; break
                            if placed: break
                            
                    if not placed:
                        for d in days:
                            if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, True, course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state, hide_c004=hide_c004):
                                h -= a; placed = True; break

                    if placed and sync_name and sync_name not in elective_sync:
                        for dcheck in days:
                            slots_used = [s_ for s_ in slot_keys if tt.at[dcheck, s_].startswith(code)]
                            if slots_used:
                                accum = []; acc_dur = 0.0
                                for s_ in slots_used:
                                    accum.append(s_); acc_dur += slot_dur[s_]
                                    if acc_dur + 1e-9 >= a:
                                        elective_sync[sync_name] = {"day": dcheck, "slots": accum.copy()}
                                        break
                                if sync_name in elective_sync: break

                    attempts += 1
                if h > 1e-9:
                    failed.append({"Label": label, "Course_Code": code, "Type": typ, "Hours_Remaining": round(h, 2), "Faculty": f})
            placed_list.append(c)
        return placed_list

    start_idx_ref = [seed % len(days)]
    elec_final.sort(key=lambda x: 0 if x.get("_sync_name") in elective_sync else 1)
    
    priority_placed = place_course_list(elec_final, start_idx_ref)
    combined_placed = assign_combined_precise_durations(tt, busy, rm, room_busy, labsd, course_usage, combined_core, rr_state=rr_state, hide_c004=hide_c004)
    regular_placed = place_course_list(regular_core, start_idx_ref)

    ws.append(["Day"] + slot_keys)
    for d in days:
        ws.append([d] + [tt.at[d, s] for s in slot_keys])
    ws.append([""])
    return (priority_placed + regular_placed + combined_core), failed

def split(c):
    f = [x for x in c if s(x.get("Semester_Half","")) in ["1","0"]]
    s2 = [x for x in c if s(x.get("Semester_Half","")) in ["2","0"]]
    return f, s2

def render_excel_to_html(ws):
    # Reduced outer padding slightly and added box-sizing to prevent overflow
    html = '<div style="width: 100%; overflow-x: auto; padding: 12px; background-color: #1a1a1a; border-radius: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.4); box-sizing: border-box;">'
    
    # CHANGED: width: 100% and table-layout: fixed to remove the horizontal scrollbar. 
    # Reduced base font-size to 12px for the "zoomed out" effect.
    html += '<table style="border-collapse: collapse; width: 100%; table-layout: fixed; font-family: \'Inter\', \'Segoe UI\', Roboto, Helvetica, sans-serif; font-size: 12px; color: #ffffff;">'
    
    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_cells_map[(min_row, min_col)] = {
            'colspan': max_col - min_col + 1,
            'rowspan': max_row - min_row + 1
        }
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                merged_cells_map[(r, c)] = 'skip'
    
    for row in ws.iter_rows():
        html += '<tr>'
        for cell in row:
            r, c = cell.row, cell.column
            if merged_cells_map.get((r, c)) == 'skip':
                continue
            
            colspan = merged_cells_map.get((r, c), {}).get('colspan', 1) if isinstance(merged_cells_map.get((r, c)), dict) else 1
            rowspan = merged_cells_map.get((r, c), {}).get('rowspan', 1) if isinstance(merged_cells_map.get((r, c)), dict) else 1
            
            val = cell.value if cell.value is not None else ""
            val_str = str(val).replace('\n', '<br>')
            
            # CHANGED: white-space: normal and word-wrap: break-word allow text to wrap inside the blocks!
            # Decreased padding to fit more text.
            styles = [
                "padding: 6px 4px", 
                "white-space: normal", 
                "word-wrap: break-word", 
                "line-height: 1.3"
            ]
            
            # Subtle grid lines
            if cell.border and cell.border.top and cell.border.top.style:
                styles.append("border: 1px solid #444444")
            else:
                styles.append("border: 1px solid #333333") 
                
            # Exact Fill Color Match
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != "00000000":
                color = cell.fill.start_color.rgb
                if isinstance(color, str):
                    color = "#" + color[-6:]
                    # Darker text shadow to make white text highly visible on light backgrounds
                    styles.append(f"background-color: {color}; color: #ffffff; text-shadow: 0px 1px 3px rgba(0,0,0,0.8);")
                    
            # Exact Font Match
            if cell.font:
                if cell.font.bold: styles.append("font-weight: 600")
                if cell.font.size: 
                    # Scale down the font size from Excel slightly to match the zoomed out look
                    styles.append(f"font-size: {max(9, cell.font.size - 1.5)}pt") 
                if cell.font.color and cell.font.color.rgb and cell.font.color.rgb != "00000000":
                    fcolor = cell.font.color.rgb
                    if isinstance(fcolor, str) and fcolor != "FF000000": 
                        styles.append(f"color: #{fcolor[-6:]}")
                    
            # Exact Alignment Match
            if cell.alignment:
                halign = cell.alignment.horizontal if cell.alignment.horizontal else "center"
                styles.append(f"text-align: {halign}")
                valign = cell.alignment.vertical if cell.alignment.vertical else "middle"
                styles.append(f"vertical-align: {valign}")
            else:
                styles.append("text-align: center; vertical-align: middle;")
                
            html += f'<td colspan="{colspan}" rowspan="{rowspan}" style="{"; ".join(styles)}">{val_str}</td>'
        html += '</tr>'
    html += '</table></div>'
    return html
# ==========================================
# 4. STREAMLIT UI COMPONENTS
# ==========================================

st.title("📅 Automated College Timetable Generator")
st.markdown("Upload your required configuration CSV and JSON files. The generator will safely ignore files you don't upload (returning empty timetables for those sections).")

col1, col2 = st.columns(2)
with col1:
    st.subheader("1. Essential Configs")
    time_slots_file = st.file_uploader("Upload time_slots.json", type=['json'])
    rooms_file = st.file_uploader("Upload rooms.csv", type=['csv'])
    reg_file = st.file_uploader("Upload registrations.csv (Optional)", type=['csv'])

with col2:
    st.subheader("2. Course Data")
    course_files = st.file_uploader("Upload all your Course CSVs", type=['csv'], accept_multiple_files=True)
    seed_input = st.number_input("Random Seed (Leave blank for random)", value=None, placeholder="42", step=1)

with st.expander("ℹ️ Expected Course CSV Filenames for accurate mapping"):
    st.markdown("""
    Your files should strictly be named as follows to be routed to the correct schedule tabs:
    * `coursesCSEA-I.csv`, `coursesCSEB-I.csv`
    * `coursesCSEA-III.csv`, `coursesCSEB-III.csv`
    * `coursesCSE-V.csv`
    * `coursesDSAI-I.csv`, `coursesDSAI-III.csv`, `coursesDSAI-V.csv`
    * `coursesECE-I.csv`, `coursesECE-III.csv`, `coursesECE-V.csv`
    * `courses7.csv`
    """)

if st.button("🚀 Generate Timetable", type="primary"):
    if not time_slots_file or not rooms_file:
        st.error("Please ensure you have uploaded both `time_slots.json` and `rooms.csv`.")
    else:
        with st.spinner("Crunching slots..."):
            try:
                # ---------------------------------------------
                # 1. Update Global Parameters with Uploaded Data
                # ---------------------------------------------
                seed = int(seed_input) if seed_input is not None else random.randint(0, 999999)
                random.seed(seed)
                
                color_avail = colors.copy()
                random.shuffle(color_avail)
                color_map = {}
                
                # Parse Time Slots
                time_slots_file.seek(0)
                slots_data = json.load(time_slots_file)["time_slots"]
                sn = [{"key": f"{s['start']}-{s['end']}", "start": s['start'], "end": s['end'], "dur": (t2m(s["end"]) - t2m(s["start"])) / 60.0} for s in slots_data]
                sn.sort(key=lambda x: t2m(x["start"]))
                slot_keys = [s["key"] for s in sn]
                slot_dur = {s["key"]: s["dur"] for s in sn}
                
                # Parse Rooms
                rooms_file.seek(0)
                r_df = pd.read_csv(rooms_file)
                r_df["Room_ID"] = r_df["Room_ID"].astype(str).str.strip()
                cls = r_df[r_df["Room_ID"].str.startswith('C')].copy()
                labs = r_df[r_df["Room_ID"].str.startswith('L')].copy()
                
                # Parse Registrations
                if reg_file:
                    reg_file.seek(0)
                    reg = pd.read_csv(reg_file)
                    reg.set_index("Course_Code", inplace=True)
                else:
                    reg = None

                # Process all course data
                dfs = {}
                courses_dict = {}
                for f in course_files:
                    f.seek(0)
                    df = pd.read_csv(f)
                    fname = f.name
                    dfs[fname] = df
                    courses_dict[fname] = df.to_dict(orient="records")

                # Helper to safely retrieve list / DF
                def get_cdata(fname): return courses_dict.get(fname, [])
                def get_cdf(fname): return dfs.get(fname, pd.DataFrame())

                # ---------------------------------------------
                # 2. Excel Generation Logic
                # ---------------------------------------------
                wb = Workbook()
                elective_room_map = {}
                global_room_busy = {d: {} for d in days}
                sync_sem1, sync_sem3, sync_sem5, sync_sem7 = {}, {}, {}, {}
                reports = []

                # --- CSE-I ---
                ws1 = wb.active
                ws1.title = "CSE-I Timetable"
                cAf, cAs = split(get_cdata("coursesCSEA-I.csv"))
                cBf, cBs = split(get_cdata("coursesCSEB-I.csv"))
                
                csea_b1, csea_f1 = generate(cAf, ws1, "CSEA I First Half", seed+0, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy, hide_c004=True)
                csea_b2, csea_f2 = generate(cAs, ws1, "CSEA I Second Half", seed+1, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy, hide_c004=True)
                reports.extend(csea_f1 + csea_f2)
                add_csv_legend_block(ws1, get_cdf("coursesCSEA-I.csv"), "CSEA I", room_prefix="C1", elective_room_map=elective_room_map)
                
                cseb_b1, cseb_f1 = generate(cBf, ws1, "CSEB I First Half", seed+2, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy, hide_c004=True)
                cseb_b2, cseb_f2 = generate(cBs, ws1, "CSEB I Second Half", seed+3, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy, hide_c004=True)
                reports.extend(cseb_f1 + cseb_f2)
                add_csv_legend_block(ws1, get_cdf("coursesCSEB-I.csv"), "CSEB I", room_prefix="C1", elective_room_map=elective_room_map)
                
                merge_and_color(ws1, (csea_b1 or []) + (csea_b2 or []) + (cseb_b1 or []) + (cseb_b2 or []))

                # --- DSAI-I ---
                ws7 = wb.create_sheet("DSAI-I Timetable")
                d1f_i, d1s_i = split(get_cdata("coursesDSAI-I.csv"))
                dsai1_b1, dsai1_f1 = generate(d1f_i, ws7, "DSAI-I First Half", seed+16, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                dsai1_b2, dsai1_f2 = generate(d1s_i, ws7, "DSAI-I Second Half", seed+17, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(dsai1_f1 + dsai1_f2)
                add_csv_legend_block(ws7, get_cdf("coursesDSAI-I.csv"), "DSAI I", room_prefix="C1", elective_room_map=elective_room_map)
                merge_and_color(ws7, (dsai1_b1 or []) + (dsai1_b2 or []))

                # --- ECE-I ---
                ws9 = wb.create_sheet("ECE-I Timetable")
                e1f_i, e1s_i = split(get_cdata("coursesECE-I.csv"))
                ece1_b1, ece1_f1 = generate(e1f_i, ws9, "ECE-I First Half", seed+20, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                ece1_b2, ece1_f2 = generate(e1s_i, ws9, "ECE-I Second Half", seed+21, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(ece1_f1 + ece1_f2)
                add_csv_legend_block(ws9, get_cdf("coursesECE-I.csv"), "ECE I", room_prefix="C4", elective_room_map=elective_room_map)
                merge_and_color(ws9, (ece1_b1 or []) + (ece1_b2 or []))

                # --- CSE-III ---
                ws2 = wb.create_sheet("CSE-III Timetable")
                c1f, c1s = split(get_cdata("coursesCSEA-III.csv"))
                c2f, c2s = split(get_cdata("coursesCSEB-III.csv"))
                
                csea3_b1, csea3_f1 = generate(c1f, ws2, "CSEA III First Half", seed+4, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                csea3_b2, csea3_f2 = generate(c1s, ws2, "CSEA III Second Half", seed+5, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(csea3_f1 + csea3_f2)
                add_csv_legend_block(ws2, get_cdf("coursesCSEA-III.csv"), "CSEA III", room_prefix="C2", elective_room_map=elective_room_map)
                
                cseb3_b1, cseb3_f1 = generate(c2f, ws2, "CSEB III First Half", seed+6, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                cseb3_b2, cseb3_f2 = generate(c2s, ws2, "CSEB III Second Half", seed+7, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(cseb3_f1 + cseb3_f2)
                add_csv_legend_block(ws2, get_cdf("coursesCSEB-III.csv"), "CSEB III", room_prefix="C2", elective_room_map=elective_room_map)
                
                merge_and_color(ws2, (csea3_b1 or []) + (csea3_b2 or []) + (cseb3_b1 or []) + (cseb3_b2 or []))

                # --- DSAI-III ---
                ws4 = wb.create_sheet("DSAI-III Timetable")
                d1f, d1s = split(get_cdata("coursesDSAI-III.csv"))
                dsa_b1, dsa_f1 = generate(d1f, ws4, "DSAI-III First Half", seed+10, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                dsa_b2, dsa_f2 = generate(d1s, ws4, "DSAI-III Second Half", seed+11, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(dsa_f1 + dsa_f2)
                add_csv_legend_block(ws4, get_cdf("coursesDSAI-III.csv"), "DSAI", room_prefix="C4", elective_room_map=elective_room_map)
                merge_and_color(ws4, (dsa_b1 or []) + (dsa_b2 or []))

                # --- ECE-III ---
                ws5 = wb.create_sheet("ECE-III Timetable")
                e1f, e1s = split(get_cdata("coursesECE-III.csv"))
                ece_b1, ece_f1 = generate(e1f, ws5, "ECE-III First Half", seed+12, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                ece_b2, ece_f2 = generate(e1s, ws5, "ECE-III Second Half", seed+13, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(ece_f1 + ece_f2)
                add_csv_legend_block(ws5, get_cdf("coursesECE-III.csv"), "ECE", room_prefix="C4", elective_room_map=elective_room_map)
                merge_and_color(ws5, (ece_b1 or []) + (ece_b2 or []))

                # --- CSE-V ---
                ws3 = wb.create_sheet("CSE-V Timetable")
                c5f, c5s = split(get_cdata("coursesCSE-V.csv"))
                c5_b1, c5_f1 = generate(c5f, ws3, "CSE-V First Half", seed+8, sync_sem5, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                c5_b2, c5_f2 = generate(c5s, ws3, "CSE-V Second Half", seed+9, sync_sem5, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(c5_f1 + c5_f2)
                add_csv_legend_block(ws3, get_cdf("coursesCSE-V.csv"), "CSE V", room_prefix="C3", elective_room_map=elective_room_map)
                merge_and_color(ws3, (c5_b1 or []) + (c5_b2 or []))

                # --- DSAI-V ---
                ws8 = wb.create_sheet("DSAI-V Timetable")
                d5f_v, d5s_v = split(get_cdata("coursesDSAI-V.csv"))
                dsai5_b1, dsai5_f1 = generate(d5f_v, ws8, "DSAI-V First Half", seed+18, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                dsai5_b2, dsai5_f2 = generate(d5s_v, ws8, "DSAI-V Second Half", seed+19, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(dsai5_f1 + dsai5_f2)
                add_csv_legend_block(ws8, get_cdf("coursesDSAI-V.csv"), "DSAI V", room_prefix="C4", elective_room_map=elective_room_map)
                merge_and_color(ws8, (dsai5_b1 or []) + (dsai5_b2 or []))

                # --- ECE-V ---
                ws10 = wb.create_sheet("ECE-V Timetable")
                e5f_v, e5s_v = split(get_cdata("coursesECE-V.csv"))
                ece5_b1, ece5_f1 = generate(e5f_v, ws10, "ECE-V First Half", seed+22, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                ece5_b2, ece5_f2 = generate(e5s_v, ws10, "ECE-V Second Half", seed+23, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(ece5_f1 + ece5_f2)
                add_csv_legend_block(ws10, get_cdf("coursesECE-V.csv"), "ECE V", room_prefix="C4", elective_room_map=elective_room_map)
                merge_and_color(ws10, (ece5_b1 or []) + (ece5_b2 or []))

                # --- 7th Sem ---
                ws6 = wb.create_sheet("7TH-SEM Timetable")
                s7f, s7s = split(get_cdata("courses7.csv"))
                s7_b1, s7_f1 = generate(s7f, ws6, "7TH-SEM First Half", seed+14, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                s7_b2, s7_f2 = generate(s7s, ws6, "7TH-SEM Second Half", seed+15, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy)
                reports.extend(s7_f1 + s7_f2)
                add_csv_legend_block(ws6, get_cdf("courses7.csv"), "7TH SEM", room_prefix="C3", elective_room_map=elective_room_map)
                merge_and_color(ws6, (s7_b1 or []) + (s7_b2 or []))

                # --- Error Reporting ---
                if reports:
                    wsr = wb.create_sheet("Report")
                    wsr.append(["Unplaced/Partial Courses"])
                    wsr.append(["Label", "Course Code", "Type", "Hours Remaining", "Faculty"])
                    for r in reports:
                        wsr.append([r.get("Label",""), r.get("Course_Code",""), r.get("Type",""), r.get("Hours_Remaining",""), r.get("Faculty","")])
                    for col in wsr.columns:
                        maxl = 0; cl = col[0].column_letter
                        for cell in col:
                            v = cell.value
                            if v is None: continue
                            maxl = max(maxl, len(str(v)))
                        wsr.column_dimensions[cl].width = min(maxl + 2 if maxl else 8, 60)

                # Prepare the generated file for download
                excel_io = io.BytesIO()
                wb.save(excel_io)
                excel_io.seek(0)
                
                st.success("✅ Timetable successfully generated!")
                
                st.download_button(
                    label="📥 Download Excel Timetable",
                    data=excel_io,
                    file_name="Balanced_Timetable.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

                st.divider()
                st.subheader("📊 Timetable Preview")
                
                sheet_names = wb.sheetnames
                tabs = st.tabs(sheet_names)
                
                for tab, sheet_name in zip(tabs, sheet_names):
                    with tab:
                        ws_current = wb[sheet_name]
                        html_table = render_excel_to_html(ws_current)
                        # Renders inside an isolated iframe, perfectly mirroring Excel
                        components.html(html_table, height=800, scrolling=True)

            except Exception as e:
                st.error(f"❌ An error occurred during generation: {e}")