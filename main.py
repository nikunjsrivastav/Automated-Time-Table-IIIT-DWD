import pandas as pd
import json
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

with open("data/time_slots.json") as f:
    slots = json.load(f)["time_slots"]

slot_keys = [f"{slot['start'].strip()}-{slot['end'].strip()}" for slot in slots]

def slot_duration(slot):
    start, end = slot.split("-")
    h1, m1 = map(int, start.split(":"))
    h2, m2 = map(int, end.split(":"))
    return (h2 + m2 / 60) - (h1 + m1 / 60)

slot_durations = {s: slot_duration(s) for s in slot_keys}

courses = pd.read_csv("data/courses.csv").to_dict(orient="records")
rooms_df = pd.read_csv("data/rooms.csv")  # must have columns: Room_ID, Type
classrooms = rooms_df[rooms_df["Type"].str.lower() == "classroom"]["Room_ID"].tolist()
labs = rooms_df[rooms_df["Type"].str.lower() == "lab"]["Room_ID"].tolist()

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded_slots = ["07:30-09:00", "13:15-14:00", "17:30-18:30"]
MAX_ATTEMPTS = 10

colors = [
    "FFB3BA", "BAE1FF", "BAFFC9", "FFFFBA", "FFD8BA", "E3BAFF", "D0BAFF",
    "FFCBA4", "C7FFD8", "B8E1FF", "F7FFBA", "FFDFBA", "E9BAFF", "BAFFD9",
    "FFE1BA", "BAFFF2", "D1FFBA"
]

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def get_free_blocks(timetable, day):
    free_blocks = []
    block = []
    for slot in slot_keys:
        if timetable.at[day, slot] == "" and slot not in excluded_slots:
            block.append(slot)
        else:
            if block:
                free_blocks.append(block)
                block = []
    if block:
        free_blocks.append(block)
    return free_blocks

course_room_map = {}

def allocate_session(timetable, lecturer_busy, day, faculty, code, duration_hours, session_type="L", is_elective=False):
    free_blocks = get_free_blocks(timetable, day)
    for block in free_blocks:
        total = sum(slot_durations[s] for s in block)
        if total >= duration_hours:
            slots_to_use = []
            dur_accum = 0
            for s in block:
                slots_to_use.append(s)
                dur_accum += slot_durations[s]
                if dur_accum >= duration_hours:
                    break

            if not is_elective:
                if code in course_room_map:
                    room = course_room_map[code]
                else:
                    if session_type == "P":
                        if not labs:
                            print(f"No labs available for {code}")
                            return False
                        room = random.choice(labs)
                    else:
                        if not classrooms:
                            print(f"No classrooms available for {code}")
                            return False
                        room = random.choice(classrooms)
                    course_room_map[code] = room

            for s in slots_to_use:
                if session_type == "L":
                    timetable.at[day, s] = f"{code} ({room})" if not is_elective else code
                elif session_type == "T":
                    timetable.at[day, s] = f"{code}T ({room})" if not is_elective else f"{code}T"
                elif session_type == "P":
                    timetable.at[day, s] = f"{code} (Lab-{room})" if not is_elective else code

            if faculty:
                lecturer_busy[day].append(faculty)
            return True
    return False

def merge_and_style_cells(filename):
    wb = load_workbook(filename)
    ws = wb.active

    course_colors = {}
    available_colors = colors.copy()
    random.shuffle(available_colors) 

    for row_idx in range(2, ws.max_row + 1):
        start_col = 2
        while start_col <= ws.max_column:
            cell_value = ws.cell(row=row_idx, column=start_col).value
            if not cell_value:
                start_col += 1
                continue

            merge_cols = [start_col]
            if "(" in str(cell_value):
                if "Lab" in cell_value:
                    session_dur = 2
                elif "T" in cell_value:
                    session_dur = 1
                else:
                    session_dur = 1.5
            else:
                session_dur = 1.5

            total_dur = slot_durations[slot_keys[start_col-2]]
            next_col = start_col + 1
            while next_col <= ws.max_column:
                next_value = ws.cell(row=row_idx, column=next_col).value
                if next_value == cell_value:
                    total_dur += slot_durations[slot_keys[next_col-2]]
                    merge_cols.append(next_col)
                    if total_dur >= session_dur:
                        break
                    next_col += 1
                else:
                    break

            if len(merge_cols) > 1:
                ws.merge_cells(start_row=row_idx, start_column=merge_cols[0],
                               end_row=row_idx, end_column=merge_cols[-1])

            cell = ws.cell(row=row_idx, column=merge_cols[0])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

            raw_course_name = str(cell_value).split()[0].replace("T","") 
            if raw_course_name not in course_colors:
                if available_colors:
                    course_colors[raw_course_name] = available_colors.pop()
                else:
                    r = lambda: random.randint(150, 255)
                    course_colors[raw_course_name] = f"{r():02X}{r():02X}{r():02X}"
            fill_color = course_colors[raw_course_name]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            for col in merge_cols:
                ws.cell(row=row_idx, column=col).border = thin_border

            start_col = merge_cols[-1] + 1

    wb.save(filename)

def generate_timetable(courses_to_allocate, filename):
    timetable = pd.DataFrame("", index=days, columns=slot_keys)
    lecturer_busy = {day: [] for day in days}
    global course_room_map
    course_room_map = {}

    electives = [c for c in courses_to_allocate if str(c.get("Elective", 0)) == "1"]
    non_electives = [c for c in courses_to_allocate if str(c.get("Elective", 0)) != "1"]

    if electives:
        chosen_elective = random.choice(electives)
        elective_course = {
            "Course_Code": "Elective",
            "Faculty": chosen_elective.get("Faculty", ""),
            "L-T-P-S-C": chosen_elective["L-T-P-S-C"]
        }
        non_electives.append(elective_course)

    for course in non_electives:
        faculty = str(course.get("Faculty", "")).strip()
        code = str(course["Course_Code"]).strip()
        is_elective = True if code == "Elective" else False

        try:
            L, T, P, S, C = map(int, [x.strip() for x in course["L-T-P-S-C"].split("-")])
        except:
            continue

        lecture_hours_remaining = L
        attempts = 0
        while lecture_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in days:
                if lecture_hours_remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                    continue
                alloc_hours = min(1.5, lecture_hours_remaining)
                if allocate_session(timetable, lecturer_busy, day, faculty, code, alloc_hours, "L", is_elective):
                    lecture_hours_remaining -= alloc_hours
                    break

        tutorial_hours_remaining = T
        attempts = 0
        while tutorial_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in days:
                if tutorial_hours_remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                    continue
                if allocate_session(timetable, lecturer_busy, day, faculty, code, 1, "T", is_elective):
                    tutorial_hours_remaining -= 1
                    break
                    
        practical_hours_remaining = P
        attempts = 0
        while practical_hours_remaining > 0 and attempts < MAX_ATTEMPTS:
            attempts += 1
            for day in days:
                if practical_hours_remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                    continue
                alloc_hours = min(2, practical_hours_remaining) if practical_hours_remaining >= 2 else practical_hours_remaining
                if allocate_session(timetable, lecturer_busy, day, faculty, code, alloc_hours, "P", is_elective):
                    practical_hours_remaining -= alloc_hours
                    break

    for day in days:
        for slot in excluded_slots:
            if slot in timetable.columns:
                timetable.at[day, slot] = ""

    timetable.to_excel(filename, index=True)
    merge_and_style_cells(filename)
    print(f"Saved styled timetable to {filename}")

courses_first_half = [c for c in courses if str(c.get("Semester_Half")).strip() in ["1", "0"]]
courses_second_half = [c for c in courses if str(c.get("Semester_Half")).strip() in ["2", "0"]]

generate_timetable(courses_first_half, "timetable_first_half.xlsx")
generate_timetable(courses_second_half, "timetable_second_half.xlsx")

