import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import json
import glob
import re
import random

# -------------------- Styling --------------------
thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# -------------------- Time Slots --------------------
with open("data/time_slots.json") as f:
    slots = json.load(f)["time_slots"]

def t2m(x):
    h, m = map(int, x.split(":"))
    return h * 60 + m

slots_sorted = sorted(slots, key=lambda z: t2m(z["start"]))
slot_keys = [f"{s['start']}-{s['end']}" for s in slots_sorted]
slot_dur = {
    f"{s['start']}-{s['end']}":
    (t2m(s["end"]) - t2m(s["start"])) / 60.0
    for s in slots_sorted
}

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
HALVES = ["First Half", "Second Half"]

# -------------------- Helpers --------------------
def extract_code(x):
    if x is None:
        return ""
    x = str(x).strip()
    if x == "":
        return ""
    return x.split()[0].upper()

def split_faculty(fac):
    if fac is None:
        return []
    fac = str(fac)
    return [x.strip() for x in re.split(r"[\\/;,&]| and ", fac) if x.strip()]

def add_section(cell_value, section):
    cell_value = str(cell_value)
    if "(" in cell_value and cell_value.endswith(")"):
        return cell_value
    return f"{cell_value} ({section})"

def safe_title(x):
    return re.sub(r'[:\\/*?\[\]]', '_', x)[:30]

# -------------------- Load Course Info --------------------
def load_all_course_info():
    files = glob.glob("data/*.csv")
    fac_map = {}
    p_map = {}
    elective_by_basket = {}

    for f in files:
        try:
            df = pd.read_csv(f)
        except:
            continue

        cols = {c.lower(): c for c in df.columns}
        code_col = cols.get("course_code") or cols.get("course code")
        fac_col = cols.get("faculty")
        ltp_col = cols.get("l-t-p-s-c") or cols.get("l_t_p_s_c") or cols.get("ltp")
        elec_col = cols.get("elective")
        basket_col = cols.get("electivebasket") or cols.get("elective_basket")

        if not code_col:
            continue

        for _, r in df.iterrows():
            c = str(r.get(code_col, "")).strip()
            if c == "" or c.lower() in {"nan", "new"}:
                continue

            c = c.upper()
            facs = split_faculty(r.get(fac_col, ""))
            fac_map[c] = facs

            try:
                parts = [float(x) for x in re.split(r"[-:]", str(r.get(ltp_col, ""))) if x]
                P = parts[2] if len(parts) >= 3 else 0.0
            except:
                P = 0.0

            p_map[c] = P

            elect = str(r.get(elec_col, "")).strip().lower()
            basket = str(r.get(basket_col, "")).strip()

            if elect in {"1", "yes"}:
                key = basket if basket and basket != "0" else c
                elective_by_basket.setdefault(key, []).append(c)

    return fac_map, p_map, elective_by_basket

course_faculty, course_P, elective_baskets = load_all_course_info()

# -------------------- Colors --------------------
colors = [
    "FFB3BA","BAE1FF","BAFFC9","FFFFBA","FFD8BA","E3BAFF","D0BAFF","FFCBA4",
    "C7FFD8","B8E1FF","F7FFBA","FFDFBA","E9BAFF","BAFFD9","FFE1BA","BAFFF2"
]
random.shuffle(colors)

color_map = {}
fill_map = {}

def get_fill(code):
    if not code:
        return None
    if code not in color_map:
        color_map[code] = colors.pop() if colors else "CCCCCC"
    col = color_map[code]
    if col not in fill_map:
        fill_map[col] = PatternFill(start_color=col, end_color=col, fill_type="solid")
    return fill_map[col]

# -------------------- Read Input Timetable --------------------
wb_in = openpyxl.load_workbook("Balanced_Timetable_latest.xlsx")
faculty_slots = {}

for sheet in wb_in.sheetnames:
    ws = wb_in[sheet]
    current_half = None

    rows = list(ws.values)
    if not rows:
        continue

    header_index = None
    for i, row in enumerate(rows):
        if row and row[0] == "Day":
            header_index = i
            break

    if header_index is None:
        continue

    header = rows[header_index]
    col_map = {name: idx for idx, name in enumerate(header) if name in slot_keys}

    for r in range(1, ws.max_row + 1):
        first_cell = str(ws.cell(r, 1).value).strip()

        if "First Half" in first_cell:
            current_half = "First Half"
            continue
        if "Second Half" in first_cell:
            current_half = "Second Half"
            continue
        if first_cell not in days or not current_half:
            continue

        day = first_cell

        for sk, col in col_map.items():
            cell = ws.cell(r, col + 1).value
            if not cell:
                continue
            code = extract_code(cell)

# ---------- ELECTIVES ----------
            if code.startswith("ELECTIVE"):
                m = re.search(r"ELECTIVE\s*([0-9]+)", code)
                val_with_section = cell  # keep cell text as-is

    # Case 1: Elective with basket number
                if m:
                    basket = m.group(1)
                    reps = elective_baskets.get(basket, [])
                else:
        # fallback: all electives
                    reps = [c for lst in elective_baskets.values() for c in lst]

                for c_up in reps:
                    facs = course_faculty.get(c_up, [])
                    for fac in facs:
                        if not fac:
                            continue
                        faculty_slots.setdefault(
                            fac,
                            {h: {d: {s: "" for s in slot_keys} for d in days} for h in HALVES}
                        )
                        faculty_slots[fac][current_half][day][sk] = val_with_section

# ---------- NORMAL COURSES ----------
            else:
                facs = course_faculty.get(code, [])
                for fac in facs:
                    if not fac:
                        continue
                    faculty_slots.setdefault(
                        fac,
                        {h: {d: {s: "" for s in slot_keys} for d in days} for h in HALVES}
                    )
                    faculty_slots[fac][current_half][day][sk] = cell


# -------------------- Write Output --------------------
wb_out = Workbook()
first_sheet = True

for fac, halves in faculty_slots.items():
    ws = wb_out.active if first_sheet else wb_out.create_sheet()
    ws.title = safe_title(fac)
    first_sheet = False

    for half in HALVES:
        ws.append([f"{fac} | {half}"])
        ws.append(["Day"] + slot_keys)

        table = halves[half]
        for d in days:
            ws.append([d] + [table[d][s] for s in slot_keys])

        ws.append([])

    # formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        width = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

wb_out.save("Faculty_Timetable.xlsx")
print("Faculty_Timetable.xlsx generated successfully.")
